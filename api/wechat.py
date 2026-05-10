"""
微信公众号记账机器人 - Webhook 入口
"""
import os
import io
import hmac
import hashlib
import time
import json
import re
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
from urllib.parse import unquote

from fastapi import FastAPI, Request, Response, UploadFile, File, Depends, HTTPException, status
from fastapi.responses import StreamingResponse, HTMLResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import httpx
import jwt
import secrets

app = FastAPI()

# ============ 配置（从环境变量读取）============
APPID = os.environ.get("WECHAT_APPID", "")
APPSECRET = os.environ.get("WECHAT_APPSECRET", "")
TOKEN = os.environ.get("WECHAT_TOKEN", "")
SUPABASE_URL = os.environ.get("SUPABASE_URL", "")
SUPABASE_KEY = os.environ.get("SUPABASE_KEY", "")
PUBLIC_BASE_URL = os.environ.get("PUBLIC_BASE_URL", "").rstrip("/")
REPORT_TOKEN = os.environ.get("REPORT_TOKEN", "")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "")
ADMIN_SECRET = os.environ.get("ADMIN_SECRET", secrets.token_urlsafe(32))
DEEPSEEK_API_KEY = os.environ.get("DEEPSEEK_API_KEY", "") or "sk-67752290f630459197881d12938ff2f9"
DEEPSEEK_MODEL = os.environ.get("DEEPSEEK_MODEL", "deepseek-chat")
DEEPSEEK_VISION_MODEL = os.environ.get("DEEPSEEK_VISION_MODEL", "deepseek-chat")
RETENTION_DAYS = int(os.environ.get("RETENTION_DAYS", "730"))  # 默认保存2年
ARCHIVE_BATCH = 200
EXPORT_TTL_SECONDS = 600
LOCAL_TZ = ZoneInfo("Asia/Shanghai")
UTC_TZ = ZoneInfo("UTC")
PENDING_DELETE_TTL = 300  # 秒
ALIAS_CACHE_TTL = 600
PENDING_CATEGORY_TTL = 300  # 秒
ADMIN_TOKEN_EXPIRY = 3600 * 24  # Token 24小时过期
MAX_LOGIN_ATTEMPTS = 5  # 最大登录尝试次数
LOGIN_LOCKOUT_TIME = 300  # 锁定时间（秒）

# 登录失败记录（内存，按 IP）
LOGIN_ATTEMPTS = {}

security = HTTPBearer()


def verify_admin_token(credentials: HTTPAuthorizationCredentials = Depends(security)):
    """验证管理员Token"""
    try:
        token = credentials.credentials
        payload = jwt.decode(token, ADMIN_SECRET, algorithms=["HS256"])
        if payload.get("type") != "admin":
            raise HTTPException(status_code=403, detail="Invalid token")

        timestamp = payload.get("timestamp", 0)
        if int(time.time()) - timestamp > ADMIN_TOKEN_EXPIRY:
            raise HTTPException(status_code=403, detail="Token expired")

        return payload
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=403, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=403, detail="Invalid token")


# 待确认删除（内存，按 openid）
PENDING_DELETES = {}
# 待分类选择（内存，按 openid）
PENDING_CATEGORY_PICKS = {}
# 消息去重缓存（内存，避免数据库查询）
MSG_DEDUP_CACHE = {}
MSG_DEDUP_MAX_SIZE = 1000  # 最多保留1000条消息ID
MSG_DEDUP_TTL = 300  # 消息ID保留5分钟

# ============ 分类（不再使用内置关键词，仅用用户配置的别名完全匹配）============
# 原 CATEGORY_KEYWORDS 已移除，避免未设置的类目（如交通）自动归类；未出现过的备注一律由用户选择分类。

# 关键词别名缓存（全局）
CATEGORY_ALIAS_CACHE = {"value": {}, "expires_at": 0}
# 分类列表缓存（全局）
CATEGORY_LIST_CACHE = {"value": [], "expires_at": 0}
CATEGORY_LIST_CACHE_TTL = 600  # 分类列表缓存10分钟
# 记录缓存（用于管理后台统计）
RECORDS_CACHE = {"value": [], "expires_at": 0, "count": 0}
RECORDS_CACHE_TTL = 30  # 记录缓存30秒，编辑后统计尽快更新

# ============ 数据库操作（使用 REST API）============
def get_supabase_client():
    """创建简单的 Supabase REST 客户端"""
    class SupabaseClient:
        def __init__(self, url, key):
            self.url = url.rstrip('/')
            self.key = key
            self.headers = {
                "apikey": key,
                "Authorization": f"Bearer {key}",
                "Content-Type": "application/json",
                "Prefer": "return=representation"
            }
        
        def table(self, name):
            return SupabaseTable(self.url, name, self.headers)
    
    class SupabaseTable:
        def __init__(self, base_url, name, headers):
            self.url = f"{base_url}/rest/v1/{name}"
            self.headers = headers
        
        def insert(self, data):
            class Result:
                def __init__(self, data):
                    self.data = data
                def execute(self):
                    return self
            
            response = httpx.post(self.url, json=data, headers=self.headers, timeout=8.0)
            response.raise_for_status()
            return Result(response.json() if response.content else [data])
        
        def select(self, columns="*"):
            return QueryBuilder(self.url, self.headers, columns)

        def update(self, data):
            return UpdateBuilder(self.url, self.headers, data)

        def delete(self):
            return DeleteBuilder(self.url, self.headers)
    
    class QueryBuilder:
        def __init__(self, url, headers, columns):
            self.url = url
            self.headers = headers
            self.params = {"select": columns}
            self.filters = []
        
        def eq(self, column, value):
            self.filters.append((column, "eq", value))
            return self
        
        def gte(self, column, value):
            self.filters.append((column, "gte", value))
            return self
        
        def lt(self, column, value):
            self.filters.append((column, "lt", value))
            return self

        def ilike(self, column, value):
            self.filters.append((column, "ilike", value))
            return self

        def lte(self, column, value):
            self.filters.append((column, "lte", value))
            return self
        
        def order(self, column, desc=False):
            self.params["order"] = f"{column}.{'desc' if desc else 'asc'}"
            return self

        def limit(self, count: int):
            self.params["limit"] = str(count)
            return self
        
        def execute(self):
            for column, op, value in self.filters:
                self.params[column] = f"{op}.{value}"
            
            response = httpx.get(self.url, params=self.params, headers=self.headers, timeout=8.0)
            response.raise_for_status()
            class Result:
                def __init__(self, data):
                    self.data = data
            return Result(response.json())

    class UpdateBuilder:
        def __init__(self, url, headers, data):
            self.url = url
            self.headers = headers
            self.data = data
            self.params = {}
            self.filters = []

        def eq(self, column, value):
            self.filters.append((column, "eq", value))
            return self

        def execute(self):
            for column, op, value in self.filters:
                self.params[column] = f"{op}.{value}"

            response = httpx.patch(self.url, params=self.params, json=self.data, headers=self.headers, timeout=8.0)
            response.raise_for_status()
            class Result:
                def __init__(self, data):
                    self.data = data
            return Result(response.json() if response.content else [])

    class DeleteBuilder:
        def __init__(self, url, headers):
            self.url = url
            self.headers = headers
            self.params = {}
            self.filters = []

        def eq(self, column, value):
            self.filters.append((column, "eq", value))
            return self

        def execute(self):
            for column, op, value in self.filters:
                self.params[column] = f"{op}.{value}"

            response = httpx.delete(self.url, params=self.params, headers=self.headers, timeout=8.0)
            response.raise_for_status()
            class Result:
                def __init__(self, data):
                    self.data = data
            return Result(response.json() if response.content else [])
    
    return SupabaseClient(SUPABASE_URL, SUPABASE_KEY)


def add_record(openid: str, nickname: str, amount: float, category: str, description: str, created_at: datetime = None):
    """添加记账记录"""
    try:
        # 注意：archive_old_records() 已移除，避免阻塞消息响应
        # 归档应通过定时任务或管理接口触发
        supabase = get_supabase_client()
        created_at_value = to_utc_iso(created_at or datetime.now(LOCAL_TZ))
        data = {
            "openid": openid,
            "nickname": nickname,
            "amount": amount,
            "category": category,
            "description": description,
            "created_at": created_at_value
        }
        result = supabase.table("records").insert(data).execute()
        invalidate_records_cache()  # 清除缓存
        return result
    except Exception as e:
        print(f"数据库错误: {str(e)[:100]}")
        raise


def get_records(start_date: datetime = None, end_date: datetime = None, category: str = None, limit: int = None):
    """查询记录（所有人共同）"""
    try:
        supabase = get_supabase_client()
        query = supabase.table("records").select("*")
        
        if start_date:
            query = query.gte("created_at", to_utc_iso(start_date))
        if end_date:
            query = query.lt("created_at", to_utc_iso(end_date))
        if category:
            query = query.eq("category", category)
        
        query = query.order("created_at", desc=True)
        if limit:
            query = query.limit(limit)
        result = query.execute()
        return result.data
    except Exception as e:
        print(f"查询错误: {str(e)[:100]}")
        return []


def get_records_cached(max_records: int = 5000, force_refresh: bool = False):
    """获取所有记录（带缓存，用于管理后台统计）。force_refresh=True 时强制从数据库重新加载。"""
    now = int(time.time())
    if force_refresh:
        RECORDS_CACHE["expires_at"] = 0
    if not force_refresh and RECORDS_CACHE["value"] and now < RECORDS_CACHE["expires_at"]:
        print(f"使用缓存: {len(RECORDS_CACHE['value'])} 条记录")
        return RECORDS_CACHE["value"]
    try:
        print("缓存过期或为空，从数据库加载...")
        supabase = get_supabase_client()
        query = supabase.table("records").select("*").order("created_at", desc=True).limit(max_records)
        result = query.execute()
        records = result.data
        print(f"从数据库加载了 {len(records)} 条记录")
        # 更新缓存
        RECORDS_CACHE["value"] = records
        RECORDS_CACHE["expires_at"] = now + RECORDS_CACHE_TTL
        RECORDS_CACHE["count"] = len(records)
        return records
    except Exception as e:
        import traceback
        print(f"缓存查询错误: {traceback.format_exc()}")
        # 如果有旧缓存，返回旧缓存
        if RECORDS_CACHE["value"]:
            print(f"使用旧缓存: {len(RECORDS_CACHE['value'])} 条记录")
            return RECORDS_CACHE["value"]
        return []


def invalidate_records_cache():
    """清除记录缓存（记录变动后调用）"""
    RECORDS_CACHE["expires_at"] = 0


def filter_records_by_local_range(records: list, start_date: datetime, end_date: datetime) -> list:
    """按北京时间过滤记录（左闭右开）"""
    filtered = []
    for r in records:
        dt = to_local_datetime(r["created_at"])
        if start_date <= dt < end_date:
            filtered.append(r)
    return filtered


def get_records_by_keyword(start_date: datetime = None, end_date: datetime = None, keyword: str = "", limit: int = None):
    """按描述关键词查询记录"""
    try:
        supabase = get_supabase_client()
        query = supabase.table("records").select("*")

        if start_date:
            query = query.gte("created_at", to_utc_iso(start_date))
        if end_date:
            query = query.lt("created_at", to_utc_iso(end_date))
        if keyword:
            query = query.ilike("description", f"*{keyword}*")

        query = query.order("created_at", desc=True)
        if limit:
            query = query.limit(limit)
        result = query.execute()
        return result.data
    except Exception as e:
        print(f"关键词查询错误: {str(e)[:100]}")
        return []


def get_records_by_user(openid: str, limit: int = 1):
    """获取用户最新记录"""
    try:
        supabase = get_supabase_client()
        result = (
            supabase.table("records")
            .select("*")
            .eq("openid", openid)
            .order("created_at", desc=True)
            .limit(limit)
            .execute()
        )
        return result.data
    except Exception as e:
        print(f"用户记录查询错误: {str(e)[:100]}")
        return []


def get_statistics(start_date: datetime = None, end_date: datetime = None):
    """获取统计数据（所有人共同）"""
    fetch_start = start_date - timedelta(days=1) if start_date else None
    fetch_end = end_date + timedelta(days=1) if end_date else None
    records = get_records(fetch_start, fetch_end)
    if start_date and end_date:
        records = filter_records_by_local_range(records, start_date, end_date)
    
    total = sum(r["amount"] for r in records)
    by_category = {}
    max_record = None
    
    for r in records:
        cat = r["category"]
        by_category[cat] = by_category.get(cat, 0) + r["amount"]
        if not max_record or r["amount"] > max_record["amount"]:
            max_record = r
    
    return {
        "total": total,
        "by_category": by_category,
        "count": len(records),
        "max_record": max_record,
        "latest_record": records[0] if records else None
    }


def to_local_datetime(value: str) -> datetime:
    """解析并转为北京时间"""
    dt = datetime.fromisoformat(value.replace("Z", "+00:00"))
    if dt.tzinfo is None:
        return dt.replace(tzinfo=UTC_TZ).astimezone(LOCAL_TZ)
    return dt.astimezone(LOCAL_TZ)


def to_utc_iso(dt: datetime) -> str:
    """将时间转为 UTC ISO 字符串"""
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=LOCAL_TZ)
    return dt.astimezone(UTC_TZ).isoformat()


def normalize_dash(text: str) -> str:
    """统一分隔符"""
    return (
        text.replace("–", "-")
        .replace("—", "-")
        .replace("－", "-")
        .replace("～", "-")
        .replace("~", "-")
        .replace("至", "-")
    )


def parse_date_token(token: str) -> datetime:
    """解析日期标记（支持 今天/昨天/本月/本周/MM-DD）"""
    now = datetime.now(LOCAL_TZ)
    if token in ["今天", "今日"]:
        return now.replace(hour=0, minute=0, second=0, microsecond=0)
    if token in ["昨天", "昨日"]:
        return (now - timedelta(days=1)).replace(hour=0, minute=0, second=0, microsecond=0)
    if token == "本周":
        today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        return today_start - timedelta(days=today_start.weekday())
    if token == "本月":
        return now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    if "-" in token:
        try:
            month, day = token.split("-", 1)
            month = int(month)
            day = int(day)
            dt = now.replace(month=month, day=day, hour=0, minute=0, second=0, microsecond=0)
            if dt > now:
                dt = dt.replace(year=dt.year - 1)
            return dt
        except Exception:
            return None
    return None


def update_record(record_id: int, amount: float, category: str, description: str):
    """更新记账记录"""
    supabase = get_supabase_client()
    data = {
        "amount": amount,
        "category": category,
        "description": description
    }
    result = supabase.table("records").update(data).eq("id", record_id).execute()
    invalidate_records_cache()  # 清除缓存
    return result


def delete_record(record_id: int):
    """删除记账记录，返回 True/False 表示是否成功"""
    supabase = get_supabase_client()
    result = supabase.table("records").delete().eq("id", record_id).execute()
    invalidate_records_cache()
    # 验证删除是否真正生效
    check = supabase.table("records").select("id").eq("id", record_id).limit(1).execute()
    if check.data:
        return False  # 记录仍然存在，删除被 RLS 阻止
    return True


def archive_deleted_record(record: dict, deleted_by: str):
    """保存已删除记录到回收站"""
    supabase = get_supabase_client()
    data = {
        "original_id": record["id"],
        "deleted_by": deleted_by,
        "openid": record.get("openid", ""),
        "nickname": record.get("nickname", ""),
        "amount": record.get("amount", 0),
        "category": record.get("category", ""),
        "description": record.get("description", ""),
        "created_at": record.get("created_at", ""),
        "deleted_at": datetime.now(LOCAL_TZ).isoformat()
    }
    supabase.table("records_deleted").insert(data).execute()


def get_deleted_records(deleted_by: str, limit: int = 10):
    """获取回收站记录"""
    try:
        supabase = get_supabase_client()
        result = (
            supabase.table("records_deleted")
            .select("*")
            .eq("deleted_by", deleted_by)
            .order("deleted_at", desc=True)
            .execute()
        )
        return result.data[:limit]
    except Exception as e:
        print(f"回收站查询错误: {str(e)[:100]}")
        return []


def restore_deleted_record(deleted_by: str, index: int):
    """从回收站恢复记录"""
    supabase = get_supabase_client()
    records = get_deleted_records(deleted_by, limit=20)
    if index < 1 or index > len(records):
        return {"error": "invalid"}
    record = records[index - 1]
    insert_data = {
        "openid": record.get("openid", ""),
        "nickname": record.get("nickname", ""),
        "amount": record.get("amount", 0),
        "category": record.get("category", ""),
        "description": record.get("description", ""),
        "created_at": record.get("created_at", "")
    }
    supabase.table("records").insert(insert_data).execute()
    supabase.table("records_deleted").delete().eq("id", record["id"]).execute()
    return {"restored": record}


def get_daily_total(record_date: str):
    """获取按天汇总数据"""
    try:
        supabase = get_supabase_client()
        result = supabase.table("daily_totals").select("*").eq("record_date", record_date).execute()
        return result.data[0] if result.data else None
    except Exception as e:
        print(f"汇总查询错误: {str(e)[:100]}")
        return None


def add_daily_total(record_date: str, amount: float):
    """新增或累加日汇总"""
    supabase = get_supabase_client()
    now = datetime.now(LOCAL_TZ).isoformat()
    existing = get_daily_total(record_date)
    if existing:
        new_total = float(existing.get("total_amount", 0)) + amount
        supabase.table("daily_totals").update({
            "total_amount": new_total,
            "updated_at": now
        }).eq("record_date", record_date).execute()
        return new_total

    supabase.table("daily_totals").insert({
        "record_date": record_date,
        "total_amount": amount,
        "updated_at": now
    }).execute()
    return amount


def archive_old_records():
    """归档超过保留天数的明细，只保留金额汇总"""
    try:
        if RETENTION_DAYS <= 0:
            return 0
        supabase = get_supabase_client()
        cutoff = datetime.now(LOCAL_TZ) - timedelta(days=RETENTION_DAYS)
        records = (
            supabase.table("records")
            .select("*")
            .lte("created_at", cutoff.isoformat())
            .order("created_at", desc=False)
            .limit(ARCHIVE_BATCH)
            .execute()
            .data
        )
        if not records:
            return 0

        totals_by_date = {}
        for r in records:
            dt = to_local_datetime(r["created_at"])
            date_key = dt.strftime("%Y-%m-%d")
            totals_by_date[date_key] = totals_by_date.get(date_key, 0) + float(r["amount"])

        for date_key, amount in totals_by_date.items():
            add_daily_total(date_key, amount)

        for r in records:
            delete_record(r["id"])

        return len(records)
    except Exception as e:
        print(f"归档错误: {str(e)[:100]}")
        return 0


def get_debt(name: str):
    """获取指定人的欠款记录（我欠别人）"""
    try:
        supabase = get_supabase_client()
        result = supabase.table("debts").select("*").eq("name", name).execute()
        return result.data[0] if result.data else None
    except Exception as e:
        print(f"外债查询错误: {str(e)[:100]}")
        return None


def add_debt(name: str, amount: float, note: str = ""):
    """新增或累加欠款（我欠别人）"""
    supabase = get_supabase_client()
    now = datetime.now(LOCAL_TZ).isoformat()
    existing = get_debt(name)
    if existing:
        new_amount = float(existing.get("amount", 0)) + amount
        data = {
            "amount": new_amount,
            "status": "active",
            "updated_at": now
        }
        if note:
            data["note"] = note
        supabase.table("debts").update(data).eq("name", name).execute()
        return new_amount

    data = {
        "name": name,
        "amount": amount,
        "status": "active",
        "note": note,
        "created_at": now,
        "updated_at": now
    }
    supabase.table("debts").insert(data).execute()
    return amount


def repay_debt(name: str, amount: float):
    """还钱扣减欠款（我欠别人）"""
    supabase = get_supabase_client()
    now = datetime.now(LOCAL_TZ).isoformat()
    existing = get_debt(name)
    if not existing:
        return {"error": "not_found"}

    balance = float(existing.get("amount", 0))
    if amount > balance:
        return {"error": "overpay", "balance": balance}

    new_balance = balance - amount
    status = "paid" if new_balance == 0 else "active"
    data = {
        "amount": new_balance,
        "status": status,
        "updated_at": now
    }
    supabase.table("debts").update(data).eq("name", name).execute()
    return {"balance": new_balance, "status": status}


def list_debts():
    """列出所有未清欠款（我欠别人）"""
    try:
        supabase = get_supabase_client()
        result = supabase.table("debts").select("*").eq("status", "active").order("amount", desc=True).execute()
        return result.data
    except Exception as e:
        print(f"外债列表错误: {str(e)[:100]}")
        return []


def list_debts_all(include_paid: bool = False):
    """列出外债（管理后台用，可选含已还清）"""
    try:
        supabase = get_supabase_client()
        query = supabase.table("debts").select("*").order("updated_at", desc=True)
        if not include_paid:
            query = query.eq("status", "active")
        result = query.execute()
        return result.data or []
    except Exception as e:
        print(f"外债列表错误: {str(e)[:100]}")
        return []


def delete_debt(name: str):
    """删除/清空某条外债记录"""
    try:
        supabase = get_supabase_client()
        supabase.table("debts").delete().eq("name", name).execute()
        return True
    except Exception as e:
        print(f"外债删除错误: {str(e)[:100]}")
        return False


def get_subscription(openid: str, report_type: str):
    """获取订阅记录"""
    try:
        supabase = get_supabase_client()
        result = (
            supabase.table("report_subscriptions")
            .select("*")
            .eq("openid", openid)
            .eq("report_type", report_type)
            .execute()
        )
        return result.data[0] if result.data else None
    except Exception as e:
        print(f"订阅查询错误: {str(e)[:100]}")
        return None


def subscribe_report(openid: str, report_type: str):
    """订阅周报/月报"""
    supabase = get_supabase_client()
    now = datetime.now(LOCAL_TZ).isoformat()
    existing = get_subscription(openid, report_type)
    if existing:
        return True
    supabase.table("report_subscriptions").insert({
        "openid": openid,
        "report_type": report_type,
        "created_at": now
    }).execute()
    return True


def unsubscribe_report(openid: str, report_type: str):
    """取消订阅"""
    supabase = get_supabase_client()
    supabase.table("report_subscriptions").delete().eq("openid", openid).eq("report_type", report_type).execute()


def list_subscribers(report_type: str):
    """获取订阅用户列表"""
    try:
        supabase = get_supabase_client()
        result = (
            supabase.table("report_subscriptions")
            .select("*")
            .eq("report_type", report_type)
            .execute()
        )
        return [r["openid"] for r in result.data]
    except Exception as e:
        print(f"订阅列表错误: {str(e)[:100]}")
        return []


def is_duplicate_message(msg_id: str) -> bool:
    """检查消息是否已处理（使用内存缓存，避免数据库查询延迟）"""
    if not msg_id:
        return False
    now = time.time()
    # 检查内存缓存
    if msg_id in MSG_DEDUP_CACHE:
        return True
    return False


def record_message_id(msg_id: str) -> None:
    """记录消息 ID 用于去重（使用内存缓存）"""
    if not msg_id:
        return
    now = time.time()
    MSG_DEDUP_CACHE[msg_id] = now
    # 清理过期和超量的消息ID
    if len(MSG_DEDUP_CACHE) > MSG_DEDUP_MAX_SIZE:
        # 删除过期的
        expired = [k for k, v in MSG_DEDUP_CACHE.items() if now - v > MSG_DEDUP_TTL]
        for k in expired:
            MSG_DEDUP_CACHE.pop(k, None)
        # 如果还是太多，删除最老的
        if len(MSG_DEDUP_CACHE) > MSG_DEDUP_MAX_SIZE:
            sorted_items = sorted(MSG_DEDUP_CACHE.items(), key=lambda x: x[1])
            for k, _ in sorted_items[:len(MSG_DEDUP_CACHE) - MSG_DEDUP_MAX_SIZE // 2]:
                MSG_DEDUP_CACHE.pop(k, None)


# ============ 消息解析 ============
def get_category_aliases() -> dict:
    """读取关键词别名（带缓存）"""
    now = int(time.time())
    if CATEGORY_ALIAS_CACHE["value"] and now < CATEGORY_ALIAS_CACHE["expires_at"]:
        return CATEGORY_ALIAS_CACHE["value"]
    try:
        supabase = get_supabase_client()
        result = supabase.table("category_aliases").select("keyword,category,enabled").execute()
        aliases = {}
        for row in result.data:
            if row.get("enabled", True):
                keyword = str(row.get("keyword", "")).strip().lower()
                category = str(row.get("category", "")).strip()
                if keyword and category:
                    aliases[keyword] = category
        CATEGORY_ALIAS_CACHE["value"] = aliases
        CATEGORY_ALIAS_CACHE["expires_at"] = now + ALIAS_CACHE_TTL
        return aliases
    except Exception:
        return {}


def add_category_alias(keyword: str, category: str) -> bool:
    """新增或更新关键词别名。过短(≤1字)不学习，避免误触（如「虾」误学成会员类）。"""
    keyword = keyword.strip().lower()
    category = category.strip()
    if not keyword or not category or len(keyword) <= 1:
        return False
    
    # 立即更新内存缓存，确保下次记账能匹配到
    if CATEGORY_ALIAS_CACHE["value"]:
        CATEGORY_ALIAS_CACHE["value"][keyword] = category
    
    # 异步写入数据库（不阻塞响应）
    try:
        supabase = get_supabase_client()
        existing = supabase.table("category_aliases").select("keyword").eq("keyword", keyword).limit(1).execute()
        now = datetime.now(LOCAL_TZ).isoformat()
        if existing.data:
            supabase.table("category_aliases").update({
                "category": category,
                "enabled": True,
                "updated_at": now
            }).eq("keyword", keyword).execute()
        else:
            supabase.table("category_aliases").insert({
                "keyword": keyword,
                "category": category,
                "enabled": True,
                "created_at": now,
                "updated_at": now
            }).execute()
        return True
    except Exception:
        # 数据库写入失败不影响，内存缓存已更新
        return True


def parse_category(text: str) -> str:
    """从文本识别分类：仅用用户配置的别名，完全匹配；未匹配返回空字符串（由调用方决定是否用「其他」或让用户选择）。"""
    return match_alias_category(text)


def match_alias_category(text: str) -> str:
    """仅当用户配置的别名与备注完全一致 **且目标分类仍然存在** 时才自动归类；
    否则返回空，必须让用户选择分类。不再使用任何内置关键词（如交通、餐饮）。"""
    text_lower = (text or "").strip().lower()
    if not text_lower:
        return ""
    aliases = get_category_aliases()
    matched = aliases.get(text_lower, "")
    if not matched:
        return ""
    # 验证目标分类仍然存在于用户的分类列表中（防止别名指向已删除/不存在的分类如「购物」「交通」）
    all_cats = get_all_categories()
    for cat in all_cats:
        if cat == matched or cat.startswith(matched + "|") or matched.startswith(cat + "|"):
            return matched
    # 目标分类不存在，视为无效别名
    return ""


def get_category_candidates() -> list:
    """可选分类列表（仅来自数据库/预设，不再使用内置关键词）"""
    categories = get_all_categories()
    if not categories:
        categories = ["其他"]
    elif "其他" not in categories:
        categories = list(categories) + ["其他"]
    return categories


def ai_classify(description: str, categories: list) -> str:
    """使用 DeepSeek AI 推断分类（旧接口兼容）。返回分类名或空字符串。"""
    result = ai_smart_classify(description)
    if result:
        return result
    return ""


def _call_deepseek(prompt: str, max_tokens: int = 60) -> str:
    """调用 DeepSeek API，返回文本结果"""
    if not DEEPSEEK_API_KEY:
        return ""
    response = httpx.post(
        "https://api.deepseek.com/chat/completions",
        headers={
            "Authorization": f"Bearer {DEEPSEEK_API_KEY}",
            "Content-Type": "application/json"
        },
        json={
            "model": DEEPSEEK_MODEL,
            "messages": [{"role": "user", "content": prompt}],
            "max_tokens": max_tokens,
            "temperature": 0.1
        },
        timeout=8.0
    )
    response.raise_for_status()
    data = response.json()
    return data.get("choices", [{}])[0].get("message", {}).get("content", "").strip()


def ai_batch_classify(descriptions: list) -> dict:
    """一次性为多条记录分类，返回 {描述: 分类} 的映射。"""
    if not DEEPSEEK_API_KEY or not descriptions:
        return {}

    all_cats = get_all_categories()
    cats_text = "、".join(all_cats[:30]) if all_cats else "暂无"

    items_text = "\n".join(f"{i+1}. {d}" for i, d in enumerate(descriptions))

    prompt = f"""你是记账分类助手。请为以下消费项目分配分类。

已有分类：{cats_text}

待分类项目：
{items_text}

规则：
- 优先从已有分类中选择最合适的
- 如果没有合适的已有分类，新建一个简短的分类名（2-4个字）
- 食物类：正餐、小吃、饮品、买菜 等
- 购物类：日用品、网购、服饰 等
- 出行类：交通、加油 等
- 生活类：话费、保险、维修 等

请严格按以下格式回复（每行一个，序号.分类名）：
1.分类名
2.分类名
..."""

    try:
        result = _call_deepseek(prompt, max_tokens=500)
        mapping = {}
        for line in result.strip().splitlines():
            line = line.strip()
            if not line:
                continue
            import re as _re
            m = _re.match(r'^(\d+)[.、．]\s*(.+)$', line)
            if m:
                idx = int(m.group(1)) - 1
                cat = m.group(2).strip()
                if 0 <= idx < len(descriptions) and cat:
                    mapping[descriptions[idx]] = cat
        return mapping
    except Exception as e:
        print(f"AI批量分类失败: {str(e)[:100]}")
        return {}


def ai_parse_intent(user_msg: str, recent_records_text: str = "", all_categories_text: str = "") -> dict:
    """AI 作为主大脑理解用户自然语言，返回结构化意图。"""
    if not DEEPSEEK_API_KEY:
        return {"action": "unknown", "reply": "AI 服务未配置，请联系管理员"}
    
    prompt = f"""你是一个私人记账助手。用户跟你对话来管理日常开支。请理解用户意图并返回JSON。

用户最近5条记录：
{recent_records_text or '暂无记录'}

用户已有的分类：
{all_categories_text or '暂无分类'}

用户消息："{user_msg}"

请严格返回一个JSON（不要多余文字、不要markdown）：

【记账】用户提到花了钱/买了东西/消费：
{{"action":"record","description":"物品描述","amount":数字,"category":"你认为最合适的分类(从已有分类中选,没有合适的就自己起一个简短的)"}}

【多条记账】用户一次说了多笔消费：
{{"action":"multi_record","items":[{{"description":"描述1","amount":数字,"category":"分类"}},{{"description":"描述2","amount":数字,"category":"分类"}}]}}

【删除】用户想删掉某条记录：
{{"action":"delete","description":"要删的记录关键词"}}

【撤销】用户想撤销上一条/最近一条：
{{"action":"undo"}}

【重新归类】用户想把某个东西换个分类：
{{"action":"reclassify","description":"记录关键词","category":"目标分类(用户指定了就填,没指定留空)"}}

【新建分组】用户想创建新分类/分组：
{{"action":"create_group","group":"分组名","description":"要移动的记录关键词(可选,没有就留空)"}}

【查询统计】用户想看花了多少钱：
{{"action":"query","period":"today/yesterday/week/month/7days/30days"}}

【查看明细】用户想看具体花在哪了：
{{"action":"detail","period":"today/yesterday/week/month"}}

【查分类】用户想看某个分类的花费：
{{"action":"query_category","category":"分类名或关键词"}}

【导出】用户想导出记录：
{{"action":"export","period":"month/week/all"}}

【管理后台】用户想打开网页/后台：
{{"action":"admin"}}

【闲聊/其他】无法归类为上述任何一种：
{{"action":"chat","reply":"用简短友好的中文回复用户，你是记账助手，可以引导用户记账"}}

规则：
- 金额必须是数字。"十五"=15，"一百"=100
- 如果用户说了消费但没说金额，amount设为0
- 分类要简短（2-4个字），从已有分类中选最合适的，没有就新建
- "凉皮10"="凉皮 10元"，理解为记账
- "今天花了多少"=查询today
- "删了/删掉/不要了"=删除
- "算了撤回"=撤销"""

    try:
        result = _call_deepseek(prompt, max_tokens=300)
        result = result.strip()
        if result.startswith("```"):
            result = result.split("\n", 1)[-1].rsplit("```", 1)[0].strip()
        import json
        return json.loads(result)
    except Exception as e:
        print(f"AI意图识别失败: {str(e)[:100]}")
        return {"action": "unknown", "reply": ""}


def handle_ai_intent(openid: str, nickname: str, content: str) -> str:
    """AI 作为主大脑，理解用户自然语言并执行对应操作"""
    records = get_records_by_user(openid, limit=5)
    recent_text = ""
    if records:
        lines = []
        for r in records:
            dt = to_local_datetime(r["created_at"])
            lines.append(f"{dt.strftime('%m-%d %H:%M')} {r['description']} {float(r['amount']):.0f}元 [{r['category']}]")
        recent_text = "\n".join(lines)

    all_cats = get_all_categories()
    cats_text = "、".join(all_cats[:20]) if all_cats else ""

    intent = ai_parse_intent(content, recent_text, cats_text)
    action = intent.get("action", "chat")

    if action == "record":
        amount = intent.get("amount", 0)
        desc = intent.get("description", "").strip()
        category = intent.get("category", "").strip()
        if not desc:
            return "没听清你要记什么，能再说一次吗？"
        if not amount or amount <= 0:
            return f"「{desc}」花了多少钱？"
        if category:
            _ensure_category_in_tree(category, "")
        else:
            alias_category = match_alias_category(desc)
            category = alias_category if alias_category else "其他"
        add_category_alias(desc, category)
        add_record(openid=openid, nickname=nickname, amount=amount, category=category, description=desc)
        return f"✅ 记好了\n{desc}：{amount:.2f} 元\n分类：{category}"

    elif action == "multi_record":
        items = intent.get("items", [])
        if not items:
            return "没听清你要记什么，能再说一次吗？"
        success = 0
        results = []
        for item in items:
            desc = item.get("description", "").strip()
            amount = item.get("amount", 0)
            category = item.get("category", "").strip()
            if not desc or not amount or amount <= 0:
                continue
            if category:
                _ensure_category_in_tree(category, "")
            else:
                alias_category = match_alias_category(desc)
                category = alias_category if alias_category else "其他"
            add_category_alias(desc, category)
            add_record(openid=openid, nickname=nickname, amount=amount, category=category, description=desc)
            results.append(f"{desc} {amount:.2f}元 → {category}")
            success += 1
        if success == 0:
            return "没有识别到有效的记账信息，能再说一次吗？"
        msg = f"✅ 记好了 {success} 笔：\n" + "\n".join(results)
        return msg

    elif action == "delete":
        desc = intent.get("description", "").strip()
        if not desc:
            return "要删哪条？说个关键词就行"
        all_records = get_records_by_user(openid, limit=30)
        matched = [r for r in all_records if desc in r.get("description", "")]
        if not matched:
            return f"没找到包含「{desc}」的记录"
        record = matched[0]
        archive_deleted_record(record, deleted_by=openid)
        if not delete_record(record["id"]):
            return "删除失败，数据库可能有权限问题"
        return f"✅ 已删除：{record['description']} {float(record['amount']):.2f}元"

    elif action == "undo":
        if not records:
            return "没有可以撤销的记录"
        record = records[0]
        archive_deleted_record(record, deleted_by=openid)
        if not delete_record(record["id"]):
            return "撤销失败，数据库可能有权限问题"
        return f"✅ 已撤销：{record['description']} {float(record['amount']):.2f}元"

    elif action == "reclassify":
        desc = intent.get("description", "").strip()
        target_cat = intent.get("category", "").strip()
        if not desc:
            return "要重新归类哪条？说个关键词"
        all_records = get_records_by_user(openid, limit=50)
        matched = [r for r in all_records if desc in r.get("description", "")]
        if not matched:
            return f"没找到包含「{desc}」的记录"
        if target_cat:
            new_category = target_cat
            _ensure_category_in_tree(new_category, "")
        else:
            new_category = ai_smart_classify(desc)
            if not new_category:
                return f"没法确定「{desc}」该归到哪，你想放到哪个分类？"
        updated = 0
        for record in matched:
            if record["category"] != new_category:
                update_record(record["id"], float(record["amount"]), new_category, record["description"])
                updated += 1
        add_category_alias(desc, new_category)
        if updated == 0:
            return f"「{desc}」已经在「{new_category}」里了"
        return f"✅ 已将「{desc}」的 {updated} 条记录归到「{new_category}」"

    elif action == "create_group":
        group = intent.get("group", "").strip()
        desc = intent.get("description", "").strip()
        if not group:
            return "要新建什么分组？"
        _ensure_category_in_tree(group, "")
        if not desc:
            return f"✅ 已新建分组「{group}」"
        all_records = get_records_by_user(openid, limit=50)
        matched = [r for r in all_records if desc in r.get("description", "")]
        if not matched:
            return f"✅ 已新建分组「{group}」（没找到「{desc}」的记录可移动）"
        moved = 0
        for record in matched:
            update_record(record["id"], float(record["amount"]), group, record["description"])
            moved += 1
        add_category_alias(desc, group)
        return f"✅ 已新建分组「{group}」并将「{desc}」的 {moved} 条记录移入"

    elif action == "query":
        period = intent.get("period", "today")
        period_map = {"today": "today", "yesterday": "yesterday", "week": "week", "month": "month", "7days": "7days", "30days": "30days"}
        p = period_map.get(period, "today")
        try:
            start_date, end_date = get_date_range(p)
            period_names = {"today": "今日", "yesterday": "昨日", "week": "本周", "month": "本月", "7days": "近7天", "30days": "近30天"}
            stats = get_statistics(start_date, end_date)
            return format_statistics(stats, period_names.get(p, ""), start_date, end_date)
        except Exception:
            return "查询失败，请稍后再试"

    elif action == "detail":
        period = intent.get("period", "today")
        period_map_detail = {"today": "today", "yesterday": "yesterday", "week": "week", "month": "month"}
        p = period_map_detail.get(period, "today")
        try:
            start_date, end_date = get_date_range(p)
            records_list = get_records(start_date=start_date - timedelta(days=1), end_date=end_date + timedelta(days=1), limit=50)
            records_list = filter_records_by_local_range(records_list, start_date, end_date)
            if not records_list:
                return "这段时间没有记录"
            period_names = {"today": "今日", "yesterday": "昨日", "week": "本周", "month": "本月"}
            lines = [f"📝 {period_names.get(p, '')}明细："]
            total = 0
            for i, r in enumerate(records_list, 1):
                dt = to_local_datetime(r["created_at"])
                lines.append(f"{i}) {dt.strftime('%m-%d %H:%M')} {r['description']} {float(r['amount']):.2f}元 [{r['category']}]")
                total += float(r["amount"])
            lines.append(f"\n合计：{total:.2f} 元")
            return "\n".join(lines)
        except Exception:
            return "查询失败，请稍后再试"

    elif action == "query_category":
        target_category = intent.get("category", "").strip()
        if not target_category:
            return "要查哪个分类？"
        try:
            now = datetime.now(LOCAL_TZ)
            month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            cat_records = get_records(start_date=month_start, category=target_category)
            cat_records = filter_records_by_local_range(cat_records, month_start, now + timedelta(days=1))
            if not cat_records:
                keyword_records = get_records_by_keyword(start_date=month_start, keyword=target_category)
                keyword_records = filter_records_by_local_range(keyword_records, month_start, now + timedelta(days=1))
                if not keyword_records:
                    return f"本月没有「{target_category}」相关的记录"
                cat_records = keyword_records
            total = sum(float(r["amount"]) for r in cat_records)
            count = len(cat_records)
            return f"📂 本月「{target_category}」：{total:.2f} 元（{count}笔）"
        except Exception:
            return "查询失败，请稍后再试"

    elif action == "export":
        period = intent.get("period", "month")
        period_map_export = {"month": "month", "week": "week", "all": "all"}
        p = period_map_export.get(period, "month")
        try:
            export_link = build_export_link(openid, p)
            if not export_link:
                return "导出功能未配置"
            return f"📥 导出链接（10分钟有效）：\n{export_link}"
        except Exception:
            return "导出失败，请稍后再试"

    elif action == "admin":
        if PUBLIC_BASE_URL:
            return f"🌐 管理后台：\n{PUBLIC_BASE_URL}/api/admin"
        return "管理后台未配置"

    else:
        reply = intent.get("reply", "")
        if reply:
            return reply
        return "我是你的记账助手，告诉我你花了什么钱就行，比如「午饭 25」「打车回家 30」"


def ai_smart_classify(description: str) -> str:
    """智能 AI 分类：查库 → AI推断归到最近的分类 → 新建分类。
    返回最终分类名（二级路径格式如 '餐饮|买菜'），或空字符串表示失败。
    """
    if not DEEPSEEK_API_KEY:
        return ""

    tree_paths = get_category_tree_paths()
    all_cats = get_all_categories()

    if tree_paths:
        tree_info = "\n".join(f"  - {p}" for p in sorted(tree_paths))
    elif all_cats:
        tree_info = "\n".join(f"  - {c}" for c in all_cats)
    else:
        tree_info = "  （暂无分类，需要你新建）"

    prompt = (
        f"你是一个个人记账分类助手。用户记了一笔消费：「{description}」\n\n"
        f"当前已有的分类体系（用 | 分隔层级，如 '餐饮|早餐' 表示一级是餐饮，二级是早餐）：\n"
        f"{tree_info}\n\n"
        f"请帮用户把这笔消费归类。规则：\n"
        f"1. 如果已有分类中有完全匹配的二级分类（如已有 '餐饮|买菜'），直接返回该二级分类名（只返回最末级名称，如 '买菜'）\n"
        f"2. 如果已有的一级分类中有合适的父类（如 '餐饮' 适合 '买菜'），返回格式：一级分类|新二级名称（如 '餐饮|买菜'）\n"
        f"3. 如果现有分类都不合适，新建一个合理的一级分类并归入，返回格式：新一级|新二级（如 '生活|买菜'）\n"
        f"4. 二级分类名应该简短、通用（如 '买菜' 而不是 '买西红柿'）\n\n"
        f"只回复分类结果，格式为 '一级|二级' 或已有的分类名，不要解释。"
    )

    try:
        result = _call_deepseek(prompt)
        if not result or result == "不确定":
            return ""

        result = result.strip().strip("'\"「」")

        if CATEGORY_TREE_SEP in result:
            parts = result.split(CATEGORY_TREE_SEP)
            l1 = parts[0].strip()
            l2 = parts[1].strip() if len(parts) > 1 else ""
            if l1 and l2:
                _ensure_category_in_tree(l1, l2)
                return l2
            elif l1:
                _ensure_category_in_tree(l1, "")
                return l1

        if result in all_cats:
            return result

        if tree_paths:
            for p in tree_paths:
                parts = p.split(CATEGORY_TREE_SEP)
                for part in parts:
                    if part == result:
                        return result

        _ensure_category_in_tree(result, "")
        return result

    except Exception as e:
        print(f"AI智能分类错误: {str(e)[:100]}")
        return ""


def _ensure_category_in_tree(l1: str, l2: str):
    """确保分类路径存在于类目树中，不存在则自动添加"""
    if not l1:
        return
    tree_paths = get_category_tree_paths()
    if tree_paths is None:
        tree_paths = []

    l1_path = l1
    need_save = False

    if l1_path not in tree_paths:
        tree_paths.append(l1_path)
        need_save = True

    if l2:
        l2_path = f"{l1}{CATEGORY_TREE_SEP}{l2}"
        if l2_path not in tree_paths:
            tree_paths.append(l2_path)
            need_save = True

    if need_save:
        set_category_tree(tree_paths)
        presets = get_category_presets()
        final_cat = l2 if l2 else l1
        if final_cat not in presets:
            add_category_preset(final_cat)


def ai_recognize_image(image_url: str) -> list:
    """用 DeepSeek 视觉模型识别图片中的消费记录。
    返回 [{"description": "...", "amount": 12.5}, ...] 或空列表。
    """
    if not DEEPSEEK_API_KEY:
        return []
    try:
        prompt = (
            "你是一个记账助手。请识别这张图片中的消费信息。\n"
            "图片可能是：小票、账单截图、转账记录、外卖订单截图等。\n"
            "请提取每一笔消费的【描述】和【金额】。\n\n"
            "规则：\n"
            "1. 只提取实际消费金额，忽略优惠前原价\n"
            "2. 如果有多笔消费，全部列出\n"
            "3. 描述要简短（2-6个字），如：午餐、咖啡、打车\n"
            "4. 如果图片不是消费相关的，返回空\n\n"
            "严格按以下 JSON 格式返回，不要其他文字：\n"
            '[{"description": "描述", "amount": 金额}]\n'
            "如果无法识别，返回：[]"
        )
        response = httpx.post(
            "https://api.deepseek.com/chat/completions",
            headers={
                "Authorization": f"Bearer {DEEPSEEK_API_KEY}",
                "Content-Type": "application/json"
            },
            json={
                "model": DEEPSEEK_VISION_MODEL,
                "messages": [{
                    "role": "user",
                    "content": [
                        {"type": "image_url", "image_url": {"url": image_url}},
                        {"type": "text", "text": prompt}
                    ]
                }],
                "max_tokens": 300,
                "temperature": 0.1
            },
            timeout=15.0
        )
        response.raise_for_status()
        data = response.json()
        result_text = data.get("choices", [{}])[0].get("message", {}).get("content", "").strip()
        if not result_text or result_text == "[]":
            return []
        start = result_text.find("[")
        end = result_text.rfind("]") + 1
        if start >= 0 and end > start:
            result_text = result_text[start:end]
        items = json.loads(result_text)
        valid = []
        for item in items:
            desc = str(item.get("description", "")).strip()
            try:
                amount = float(item.get("amount", 0))
            except (ValueError, TypeError):
                continue
            if desc and amount > 0:
                valid.append({"description": desc, "amount": amount})
        return valid
    except Exception as e:
        print(f"图片识别错误: {str(e)[:150]}")
        return []


def handle_image_message(openid: str, nickname: str, pic_url: str) -> str:
    """处理图片消息：识别消费并自动记账"""
    if not DEEPSEEK_API_KEY:
        return "❌ AI 功能未配置，无法识别图片"

    items = ai_recognize_image(pic_url)
    if not items:
        return "🤔 未能从图片中识别到消费记录。\n支持：小票、账单截图、转账记录、外卖订单等。"

    success = 0
    results = []
    for item in items:
        desc = item["description"]
        amount = item["amount"]
        alias_category = match_alias_category(desc)
        if not alias_category:
            alias_category = ai_smart_classify(desc)
        if not alias_category:
            alias_category = "其他"
        category = alias_category
        add_category_alias(desc, category)
        try:
            add_record(
                openid=openid,
                nickname=nickname,
                amount=amount,
                category=category,
                description=desc
            )
            success += 1
            results.append(f"  {desc} {amount:.2f}元 → {category}")
        except Exception:
            results.append(f"  {desc} {amount:.2f}元 → ❌失败")

    msg = f"📷 图片识别记账完成！成功 {success}/{len(items)} 条\n"
    msg += "\n".join(results)
    return msg


def build_category_pick_prompt(description: str, amount: float, categories: list) -> str:
    """构建分类选择提示"""
    lines = [
        f"请选择分类：",
        f"{description} {amount:.2f} 元"
    ]
    for i, cat in enumerate(categories, start=1):
        lines.append(f"{i}. {cat}")
    lines.append("回复序号即可，或发送 取消")
    return "\n".join(lines)


def parse_record_text(text: str) -> dict:
    """解析记账文本，支持多种写法：早餐8块、15块咖啡、打车 22、买菜 30 西红柿 等"""
    text = text.strip()
    # 先统一去掉金额后的 块/元/块钱（保留数字）
    text_norm = re.sub(r'(\d+(?:\.\d+)?)\s*(块钱|块|元|rmb|RMB)\s*', r'\1 ', text)
    text_norm = re.sub(r'(\d+(?:\.\d+)?)(块钱|块|元|rmb|RMB)(?=\D|$)', r'\1', text_norm)
    text_norm = text_norm.strip()

    # 分类 描述 金额（三部分，手动分类）
    explicit_match = re.match(r'^(\S+)\s+(.+?)\s+(\d+(?:\.\d+)?)\s*$', text_norm)
    if explicit_match:
        category, desc, amount = explicit_match.groups()
        return {
            "type": "record",
            "amount": float(amount),
            "description": desc.strip(),
            "category": category.strip(),
            "explicit_category": True
        }

    # 描述 金额 [备注]（如：买菜 30 西红柿）
    desc_amount_note = re.match(r'^(.+?)\s+(\d+(?:\.\d+)?)\s*(.*)$', text_norm)
    if desc_amount_note:
        desc, amount, extra = desc_amount_note.groups()
        desc = desc.strip()
        extra = extra.strip()
        amount = float(amount)
        if desc and not re.match(r'^\d+(?:\.\d+)?$', desc):  # 描述不是纯数字
            description = (desc + " " + extra) if extra else desc
            return {
                "type": "record",
                "amount": amount,
                "description": description.strip(),
                "category": desc.split()[0] if desc else "",
                "explicit_category": False
            }

    # 描述 金额（两段，无备注）
    simple_match = re.match(r'^(\S+)\s+(\d+(?:\.\d+)?)\s*$', text_norm)
    if simple_match:
        desc, amount = simple_match.groups()
        return {
            "type": "record",
            "amount": float(amount),
            "description": desc.strip(),
            "category": desc.strip(),
            "explicit_category": False
        }

    # 金额 描述（如：15 咖啡、8块 早餐）
    amount_desc = re.match(r'^(\d+(?:\.\d+)?)\s+(.+)$', text_norm)
    if amount_desc:
        amount, desc = amount_desc.groups()
        return {
            "type": "record",
            "amount": float(amount),
            "description": desc.strip(),
            "category": desc.strip(),
            "explicit_category": False
        }

    # 描述*数量 金额
    qty_match = re.match(r'^(\S+)[\*\sxX](\d+)\s+(\d+(?:\.\d+)?)$', text_norm)
    if qty_match:
        desc, qty, amount = qty_match.groups()
        total = float(qty) * float(amount)
        return {
            "type": "record",
            "amount": total,
            "description": f"{desc}*{qty}",
            "category": desc.strip(),
            "explicit_category": False
        }

    # 无空格：描述+金额 或 金额+描述（如 早餐8、15咖啡）
    no_space_desc_amount = re.match(r'^(.+?)(\d+(?:\.\d+)?)\s*$', text_norm)
    if no_space_desc_amount:
        desc, amount = no_space_desc_amount.groups()
        if desc and not re.match(r'^\d+(?:\.\d+)?$', desc):
            return {
                "type": "record",
                "amount": float(amount),
                "description": desc.strip(),
                "category": desc.strip(),
                "explicit_category": False
            }
    no_space_amount_desc = re.match(r'^(\d+(?:\.\d+)?)(.+)$', text_norm)
    if no_space_amount_desc:
        amount, desc = no_space_amount_desc.groups()
        if desc.strip():
            return {
                "type": "record",
                "amount": float(amount),
                "description": desc.strip(),
                "category": desc.strip(),
                "explicit_category": False
            }

    return {"type": "unknown"}


def parse_message(content: str) -> dict:
    """解析用户消息"""
    import re
    content = content.strip()
    
    # 查询命令
    if content in ["今日", "今天"]:
        return {"type": "query", "period": "today"}
    if content in ["昨日", "昨天"]:
        return {"type": "query", "period": "yesterday"}
    if content in ["七天", "近七天"]:
        return {"type": "query", "period": "7days"}
    if content in ["半个月", "十五天", "近半个月"]:
        return {"type": "query", "period": "15days"}
    if content in ["一个月", "近一个月", "30天"]:
        return {"type": "query", "period": "30days"}
    if content in ["本周", "这周"]:
        return {"type": "query", "period": "week"}
    if content in ["本月", "这个月"]:
        return {"type": "query", "period": "month"}
    if content in ["明细", "详情", "记录"]:
        return {"type": "detail", "period": "today"}
    if content.startswith("明细 "):
        return {"type": "detail", "period": content.split(maxsplit=1)[1].strip()}
    if content in ["帮助", "help", "?"]:
        return {"type": "help"}
    if content in ["网页", "管理", "后台", "管理后台"]:
        return {"type": "admin_url"}
    if content in ["面板", "统计面板"]:
        return {"type": "dashboard"}
    if content in ["确认删", "确认删除"]:
        return {"type": "record_delete_confirm"}
    if content in ["取消删", "取消删除"]:
        return {"type": "record_delete_cancel"}
    if content in ["上次", "最近"]:
        return {"type": "last_record"}
    if content in ["撤销", "撤销上一条"]:
        return {"type": "undo_last"}
    if content == "统计":
        return {"type": "query", "period": "7days"}

    # 导出
    export_excel_match = re.match(r'^(导出excel|导出Excel|导出表格)\s*(.*)$', content)
    if export_excel_match:
        target = export_excel_match.group(2)
        return {"type": "export", "target": target.strip() if target else ""}

    export_match = re.match(r'^导出\s*(.*)$', content)
    if export_match:
        target = export_match.group(1)
        return {"type": "export", "target": target.strip() if target else ""}

    # 快捷记账
    if content.startswith("+"):
        parsed = parse_record_text(content[1:].strip())
        if parsed["type"] == "record":
            return parsed
        return {"type": "unknown"}

    # 记一笔 分类 [金额] [备注]（如：记一笔 早餐、记一笔 打车 22、记一笔 买菜 30 西红柿）
    jiyibi_match = re.match(r'^记一笔\s+(\S+)(?:\s+(\d+(?:\.\d+)?))?\s*(.*)$', content)
    if jiyibi_match:
        category_part = jiyibi_match.group(1).strip()
        amount_part = jiyibi_match.group(2)
        note_part = (jiyibi_match.group(3) or "").strip()
        if amount_part:
            amount = float(amount_part)
            description = (category_part + " " + note_part).strip() if note_part else category_part
            return {
                "type": "record",
                "amount": amount,
                "description": description,
                "category": category_part,
                "explicit_category": False
            }
        return {"type": "record_need_amount", "category": category_part}

    # 补记（昨天/日期）
    backfill_match = re.match(r'^补记\s+(\S+)\s+(.+)$', content)
    if backfill_match:
        date_token = backfill_match.group(1).strip()
        rest = backfill_match.group(2).strip()
        parsed = parse_record_text(rest)
        if parsed["type"] == "record":
            return {
                "type": "record_backfill",
                "date_token": date_token,
                "amount": parsed["amount"],
                "description": parsed["description"],
                "category": parsed["category"]
            }
        return {"type": "unknown"}

    # 按描述删除记录：删除凉皮、删掉刚才记录的凉皮
    del_by_desc_match = re.match(r'^(删除|删掉)(刚才记录的|最近的|最近记录的|刚记的)?\s*(.+?)(的记录)?$', content)
    if del_by_desc_match:
        raw_desc = del_by_desc_match.group(3).strip()
        if raw_desc and not re.match(r'^[\d\s,，\-]+$', raw_desc):
            return {"type": "delete_by_desc", "description": raw_desc}

    # 重新归类：重新归类凉皮、归类一下凉皮
    reclassify_match = re.match(r'^(重新)?归类(一下)?\s*(.+)$', content)
    if reclassify_match:
        desc = reclassify_match.group(3).strip()
        return {"type": "reclassify_by_desc", "description": desc}

    # 新建分组并归类：新建分组 夜宵 把凉皮记录在里
    create_group_match = re.match(r'^新建(一个)?分组\s+(\S+)(?:\s+把(.+?)(?:记录在里|归到里面|放进去|移过去))?$', content)
    if create_group_match:
        group_name = create_group_match.group(2).strip()
        move_desc = (create_group_match.group(3) or "").strip()
        return {"type": "create_group_move", "group": group_name, "description": move_desc}

    # 记录修改/删除
    edit_match = re.match(r'^(改|修改)\s+(\d+)\s+(.+)$', content)
    if edit_match:
        index = int(edit_match.group(2))
        rest = edit_match.group(3).strip()
        parsed = parse_record_text(rest)
        if parsed["type"] == "record":
            return {
                "type": "record_edit",
                "index": index,
                "amount": parsed["amount"],
                "description": parsed["description"],
                "category": parsed["category"]
            }
        return {"type": "unknown"}

    delete_match = re.match(r'^(删|删除)\s*(.+)$', content)
    if delete_match:
        return {"type": "record_delete", "raw": delete_match.group(2).strip()}

    if content == "回收站":
        return {"type": "deleted_list"}

    restore_match = re.match(r'^恢复\s+(\d+)$', content)
    if restore_match:
        return {"type": "restore_deleted", "index": int(restore_match.group(1))}

    # 周报/月报订阅
    if content in ["订阅周报", "订阅月报", "取消周报", "取消月报", "周报", "月报"]:
        return {"type": "report", "action": content}

    learn_match = re.match(r'^纠错\s+(\S+)\s+(\S+)$', content)
    if learn_match:
        keyword, category = learn_match.groups()
        return {"type": "category_learn", "keyword": keyword.strip(), "category": category.strip()}

    # 分类管理
    if content in ["分类列表", "所有分类", "查看分类"]:
        return {"type": "category_list"}
    
    rename_match = re.match(r'^重命名分类\s+(\S+)\s+(\S+)$', content)
    if rename_match:
        old_name, new_name = rename_match.groups()
        return {"type": "category_rename", "old_name": old_name.strip(), "new_name": new_name.strip()}

    # 外债相关（我欠别人）
    debt_add_match = re.match(r'^欠\s+(\S+)\s+(\d+(?:\.\d+)?)\s*(.*)$', content)
    if debt_add_match:
        name, amount, note = debt_add_match.groups()
        return {"type": "debt_add", "name": name, "amount": float(amount), "note": note.strip()}

    debt_repay_match = re.match(r'^还\s+(\S+)\s+(\d+(?:\.\d+)?)$', content)
    if debt_repay_match:
        name, amount = debt_repay_match.groups()
        return {"type": "debt_repay", "name": name, "amount": float(amount)}

    debt_query_match = re.match(r'^查询外债$', content)
    if debt_query_match:
        return {"type": "debt_query_all"}

    # 自定义分类查询
    if content.startswith("分类 "):
        return {"type": "query_category", "category": content.split(maxsplit=1)[1].strip()}
    if content.startswith("统计 "):
        target = content.split(maxsplit=1)[1].strip()
        if target in ["今日", "今天", "昨天", "昨日", "七天", "近七天", "半个月", "十五天", "近半个月", "一个月", "近一个月", "本周", "本月"]:
            mapping = {
                "今日": "today",
                "今天": "today",
                "昨天": "yesterday",
                "昨日": "yesterday",
                "七天": "7days",
                "近七天": "7days",
                "半个月": "15days",
                "十五天": "15days",
                "近半个月": "15days",
                "一个月": "30days",
                "近一个月": "30days",
                "本周": "week",
                "本月": "month"
            }
            return {"type": "query", "period": mapping[target]}
        month_range = parse_month_token(target)
        if month_range:
            start_date, end_date, label = month_range
            return {"type": "query_month", "start_date": start_date, "end_date": end_date, "label": label}
        return {"type": "query_category", "category": target}

    if content.endswith("统计"):
        month_token = content.replace("统计", "").strip()
        month_range = parse_month_token(month_token)
        if month_range:
            start_date, end_date, label = month_range
            return {"type": "query_month", "start_date": start_date, "end_date": end_date, "label": label}
    
    # 分类查询：仅识别用户已有分类名（记录+预设）
    for category in get_all_categories():
        if content.strip() == category:
            return {"type": "query_category", "category": category}
    return parse_record_text(content)


def get_date_range(period: str):
    """获取日期范围"""
    now = datetime.now(LOCAL_TZ)
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    
    if period == "today":
        return today_start, today_start + timedelta(days=1)
    elif period == "yesterday":
        yesterday_start = today_start - timedelta(days=1)
        return yesterday_start, today_start
    elif period == "7days":
        return now - timedelta(days=7), now
    elif period == "15days":
        return now - timedelta(days=15), now
    elif period == "30days":
        return now - timedelta(days=30), now
    elif period == "week":
        week_start = today_start - timedelta(days=today_start.weekday())
        return week_start, now
    elif period == "month":
        month_start = today_start.replace(day=1)
        return month_start, now
    elif period == "all":
        # 导出全部：从2020年1月1日到现在（覆盖所有可能的记录）
        all_start = datetime(2020, 1, 1, 0, 0, 0, tzinfo=LOCAL_TZ)
        return all_start, now + timedelta(days=1)
    return None, None


def parse_month_token(token: str):
    """解析月份（支持 1月/01月/2025年1月/2025-01/2025/01）
    
    不指定年份时的逻辑：
    - 查询"上一次完整的该月份"
    - 如果该月已经过去（且已完整结束），取今年
    - 如果该月未来或正在进行，取去年
    """
    token = token.strip().replace(" ", "")
    if not token:
        return None
    now = datetime.now(LOCAL_TZ)
    match = re.match(r'^(?:(\d{4})[年/-])?(\d{1,2})(?:月)?$', token)
    if not match:
        return None
    year_text, month_text = match.groups()
    month = int(month_text)
    if month < 1 or month > 12:
        return None
    if year_text:
        # 明确指定了年份
        year = int(year_text)
    else:
        # 没指定年份：智能判断取今年还是去年
        year = now.year
        # 如果该月份大于当前月份（未来月份），取去年
        if month > now.month:
            year -= 1
        # 如果该月份等于或小于当前月份，取今年（因为可能有历史数据）
    
    start_date = now.replace(year=year, month=month, day=1, hour=0, minute=0, second=0, microsecond=0)
    if month == 12:
        end_date = start_date.replace(year=year + 1, month=1)
    else:
        end_date = start_date.replace(month=month + 1)
    label = f"{year}年{month}月"
    return start_date, end_date, label


def resolve_record_category(parsed: dict) -> str:
    """根据描述/显式分类确定最终分类；无匹配时返回「其他」供补记/修改用（可再在网页端改）。"""
    if parsed.get("explicit_category"):
        return parsed["category"]
    return parse_category(parsed.get("description", "")) or "其他"


def format_statistics(stats: dict, period_name: str, start_date: datetime, end_date: datetime) -> str:
    """格式化统计信息"""
    if stats["count"] == 0:
        return f"📊 {period_name}暂无记录"
    
    range_text = f"{start_date.strftime('%m-%d')} ~ {end_date.strftime('%m-%d')}"
    avg = stats["total"] / stats["count"] if stats["count"] else 0
    lines = [
        f"📊 {period_name}统计（{range_text}）",
        f"💰 总支出：{stats['total']:.2f} 元",
        ""
    ]
    
    # 按分类
    if stats["by_category"]:
        top_categories = sorted(stats["by_category"].items(), key=lambda x: -x[1])
        for cat, amount in top_categories:
            lines.append(f"{cat} {amount:.2f}")
    
    return "\n".join(lines)


def format_records(records: list, limit: int = 20) -> str:
    """格式化记录列表"""
    if not records:
        return "📝 暂无记录"
    
    lines = ["📝 最近记录（共同）："]
    for i, r in enumerate(records[:limit], start=1):
        dt = to_local_datetime(r["created_at"])
        date_str = dt.strftime("%m-%d %H:%M")
        lines.append(f"{i}. {date_str} {r['description']} {r['amount']:.2f}元 [{r['category']}]")
    
    if len(records) > limit:
        lines.append(f"  ... 共 {len(records)} 条记录")
    
    return "\n".join(lines)


def build_export_signature(openid: str, period: str, ts: int) -> str:
    """生成导出链接签名"""
    payload = f"{openid}|{period}|{ts}"
    return hmac.new(TOKEN.encode("utf-8"), payload.encode("utf-8"), hashlib.sha256).hexdigest()


def build_export_link(openid: str, period: str) -> str:
    """生成导出 Excel 的临时链接"""
    if not PUBLIC_BASE_URL:
        return ""
    ts = int(time.time())
    sig = build_export_signature(openid, period, ts)
    return f"{PUBLIC_BASE_URL}/api/export?openid={openid}&period={period}&ts={ts}&sig={sig}"


def build_report_text(period_key: str, label: str) -> str:
    """生成统计文本"""
    start_date, end_date = get_date_range(period_key)
    stats = get_statistics(start_date=start_date, end_date=end_date)
    return format_statistics(stats, label, start_date, end_date)


def verify_export_signature(openid: str, period: str, ts: str, sig: str) -> bool:
    """校验导出链接签名与有效期"""
    if not openid or not period or not ts or not sig:
        return False
    try:
        ts_int = int(ts)
    except ValueError:
        return False
    if abs(int(time.time()) - ts_int) > EXPORT_TTL_SECONDS:
        return False
    expected = build_export_signature(openid, period, ts_int)
    return hmac.compare_digest(expected, sig)


def build_export_excel_bytes(records: list, start_date: datetime, end_date: datetime, limit: int = 1000) -> bytes:
    """导出 Excel（二进制）"""
    wb = Workbook()
    ws = wb.active
    ws.title = "汇总"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9")
    )

    # 期间与类目统计
    total_amount = sum(float(r["amount"]) for r in records[:limit])
    ws.append(["统计区间", f"{start_date.strftime('%Y-%m-%d')} ~ {end_date.strftime('%Y-%m-%d')}"])
    ws.append(["记录数", len(records[:limit])])
    ws.append(["总支出", round(total_amount, 2)])
    ws.append([])

    category_totals = {}
    daily_totals = {}
    for r in records[:limit]:
        dt = to_local_datetime(r["created_at"])
        day_key = f"{dt.month}.{dt.day}"
        category = r["category"]
        amount = float(r["amount"])
        category_totals[category] = category_totals.get(category, 0) + amount
        daily_totals[day_key] = daily_totals.get(day_key, 0) + amount

    ws.append(["类目统计"])
    ws.append(["类目", "金额", "占比"])
    for cell in ws[ws.max_row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    for cat, amount in sorted(category_totals.items(), key=lambda x: -x[1]):
        percent = (amount / total_amount * 100) if total_amount else 0
        ws.append([cat, round(amount, 2), f"{percent:.1f}%"])

    ws.append([])
    ws.append(["每日合计"])
    ws.append(["日期", "金额"])
    for cell in ws[ws.max_row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border
    for day, amount in sorted(daily_totals.items()):
        ws.append([day, round(amount, 2)])

    ws.freeze_panes = "A5"
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 12

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=3):
        for cell in row:
            if cell.value is None:
                continue
            if cell.row in [1, 2, 3]:
                cell.font = Font(bold=True)
            if cell.row > 4:
                cell.border = border
            if cell.column in [2] and isinstance(cell.value, (int, float)):
                cell.number_format = "0.00"

    # 明细表
    ws_detail = wb.create_sheet("明细")
    ws_detail.append(["ID", "日期", "时间", "描述", "金额", "分类"])
    for cell in ws_detail[ws_detail.max_row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    for r in records[:limit]:
        dt = to_local_datetime(r["created_at"])
        ws_detail.append([
            r["id"],
            dt.strftime("%Y-%m-%d"),
            dt.strftime("%H:%M"),
            r["description"],
            float(r["amount"]),
            r["category"]
        ])

    ws_detail.freeze_panes = "A2"
    ws_detail.column_dimensions["A"].width = 10
    ws_detail.column_dimensions["B"].width = 14
    ws_detail.column_dimensions["C"].width = 10
    ws_detail.column_dimensions["D"].width = 30
    ws_detail.column_dimensions["E"].width = 12
    ws_detail.column_dimensions["F"].width = 12

    for row in ws_detail.iter_rows(min_row=2, max_row=ws_detail.max_row, min_col=1, max_col=6):
        for cell in row:
            if cell.value is None:
                continue
            cell.border = border
            if cell.column == 5 and isinstance(cell.value, (int, float)):
                cell.number_format = "0.00"

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()


def parse_import_excel(file_bytes: bytes) -> dict:
    """解析导入的 Excel（从明细表读取）"""
    try:
        wb = load_workbook(io.BytesIO(file_bytes))
        if "明细" not in wb.sheetnames:
            return {"error": "no_detail_sheet"}
        
        ws = wb["明细"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        updates = []
        for row in rows:
            if not row or len(row) < 6:
                continue
            record_id, date_str, time_str, description, amount, category = row[:6]
            if not record_id or not isinstance(record_id, int):
                continue
            if not description or not amount or not category:
                continue
            try:
                amount = float(amount)
            except (ValueError, TypeError):
                continue
            
            # 解析日期时间
            created_at = None
            if date_str and time_str:
                try:
                    date_str_clean = str(date_str).strip()
                    time_str_clean = str(time_str).strip()
                    # 尝试解析日期时间
                    if "-" in date_str_clean:
                        # 格式：2026-01-15
                        dt_str = f"{date_str_clean} {time_str_clean}"
                        dt = datetime.strptime(dt_str, "%Y-%m-%d %H:%M")
                    elif "/" in date_str_clean:
                        # 格式：2026/01/15
                        dt_str = f"{date_str_clean.replace('/', '-')} {time_str_clean}"
                        dt = datetime.strptime(dt_str, "%Y-%m-%d %H:%M")
                    else:
                        # 尝试其他格式
                        dt = datetime.strptime(f"{date_str_clean} {time_str_clean}", "%Y-%m-%d %H:%M")
                    created_at = dt.replace(tzinfo=LOCAL_TZ)
                except (ValueError, AttributeError):
                    # 如果解析失败，保持原值（不更新时间）
                    pass
            
            update_item = {
                "id": int(record_id),
                "description": str(description).strip(),
                "amount": amount,
                "category": str(category).strip()
            }
            if created_at:
                update_item["created_at"] = to_utc_iso(created_at)
            updates.append(update_item)
        return {"updates": updates}
    except Exception as e:
        print(f"解析导入 Excel 错误: {str(e)[:100]}")
        return {"error": "parse_failed"}


def batch_update_records(updates: list) -> dict:
    """批量更新记录（支持修改日期时间）"""
    supabase = get_supabase_client()
    success = 0
    failed = []
    for upd in updates:
        try:
            update_data = {
                "description": upd["description"],
                "amount": upd["amount"],
                "category": upd["category"]
            }
            # 如果提供了新的日期时间，也更新
            if "created_at" in upd:
                update_data["created_at"] = upd["created_at"]
            
            result = supabase.table("records").update(update_data).eq("id", upd["id"]).execute()
            if result.data:
                success += 1
            else:
                failed.append(upd["id"])
        except Exception:
            failed.append(upd["id"])
    invalidate_records_cache()
    return {"success": success, "failed": failed}


def rename_category(old_name: str, new_name: str) -> dict:
    """批量重命名分类（包括历史记录）"""
    try:
        supabase = get_supabase_client()
        # 更新所有记录
        result = supabase.table("records").update({
            "category": new_name
        }).eq("category", old_name).execute()
        
        # 更新别名表
        supabase.table("category_aliases").update({
            "category": new_name
        }).eq("category", old_name).execute()
        
        # 清除缓存，保证统计与下拉框立即使用新分类名
        CATEGORY_ALIAS_CACHE["value"] = {}
        CATEGORY_ALIAS_CACHE["expires_at"] = 0
        CATEGORY_LIST_CACHE["value"] = []
        CATEGORY_LIST_CACHE["expires_at"] = 0
        invalidate_records_cache()

        return {"success": True, "count": len(result.data) if result.data else 0}
    except Exception as e:
        print(f"重命名分类错误: {str(e)[:100]}")
        return {"success": False, "error": str(e)}


# ============ 自定义设置 ============
SETTINGS_CACHE = {"value": {}, "expires_at": 0}
SETTINGS_CACHE_TTL = 300  # 设置缓存5分钟

def get_setting(key: str, default: str = "") -> str:
    """获取设置项（带缓存）"""
    now = int(time.time())
    # 检查缓存
    if SETTINGS_CACHE["value"] and now < SETTINGS_CACHE["expires_at"]:
        return SETTINGS_CACHE["value"].get(key, default)
    try:
        supabase = get_supabase_client()
        result = supabase.table("settings").select("key,value").execute()
        settings = {}
        for item in result.data:
            settings[item["key"]] = item["value"]
        SETTINGS_CACHE["value"] = settings
        SETTINGS_CACHE["expires_at"] = now + SETTINGS_CACHE_TTL
        return settings.get(key, default)
    except Exception as e:
        print(f"获取设置错误: {str(e)[:100]}")
        return default


def set_setting(key: str, value: str) -> bool:
    """设置配置项"""
    try:
        supabase = get_supabase_client()
        # 检查是否已存在
        existing = supabase.table("settings").select("key").eq("key", key).execute()
        now = datetime.now(LOCAL_TZ).isoformat()
        if existing.data:
            supabase.table("settings").update({
                "value": value,
                "updated_at": now
            }).eq("key", key).execute()
        else:
            supabase.table("settings").insert({
                "key": key,
                "value": value,
                "created_at": now,
                "updated_at": now
            }).execute()
        # 清除缓存
        SETTINGS_CACHE["expires_at"] = 0
        return True
    except Exception as e:
        print(f"设置配置错误: {str(e)[:100]}")
        return False


# ============ 三级类目树（可选）============
CATEGORY_TREE_SEP = "|"
CATEGORY_TREE_CACHE = {"paths": None, "expires_at": 0}
CATEGORY_TREE_CACHE_TTL = 60


def get_category_tree_paths() -> list:
    """返回预设类目路径列表；空或未设置时为 None（表示使用「从记录推断」的旧逻辑）"""
    now = int(time.time())
    if CATEGORY_TREE_CACHE["paths"] is not None and now < CATEGORY_TREE_CACHE["expires_at"]:
        return CATEGORY_TREE_CACHE["paths"]
    raw = get_setting("category_tree", "").strip()
    if not raw:
        CATEGORY_TREE_CACHE["paths"] = None
        CATEGORY_TREE_CACHE["expires_at"] = now + CATEGORY_TREE_CACHE_TTL
        return None
    try:
        paths = json.loads(raw)
        if not isinstance(paths, list):
            paths = None
        else:
            paths = [str(p).strip() for p in paths if str(p).strip()]
        CATEGORY_TREE_CACHE["paths"] = paths if paths else None
    except Exception:
        CATEGORY_TREE_CACHE["paths"] = None
    CATEGORY_TREE_CACHE["expires_at"] = now + CATEGORY_TREE_CACHE_TTL
    return CATEGORY_TREE_CACHE["paths"]


def paths_to_tree(paths: list) -> dict:
    """将路径列表转为树结构：{ "一级": { "二级": ["三级1","三级2"], ... }, ... }"""
    tree = {}
    for p in paths:
        p = (p or "").strip()
        if not p:
            continue
        parts = p.split(CATEGORY_TREE_SEP)
        if len(parts) >= 1 and parts[0]:
            l1 = parts[0]
            if l1 not in tree:
                tree[l1] = {}
            if len(parts) >= 2 and parts[1]:
                l2 = parts[1]
                if l2 not in tree[l1]:
                    tree[l1][l2] = []
                if len(parts) >= 3 and parts[2]:
                    l3 = parts[2]
                    if l3 not in tree[l1][l2]:
                        tree[l1][l2].append(l3)
    return tree


def set_category_tree(paths: list) -> bool:
    """保存类目树路径列表；传入空列表则关闭「仅允许预设类目」"""
    paths = [str(p).strip() for p in paths if str(p).strip()]
    ok = set_setting("category_tree", json.dumps(paths, ensure_ascii=False))
    if ok:
        CATEGORY_TREE_CACHE["paths"] = paths if paths else None
        CATEGORY_TREE_CACHE["expires_at"] = 0
        CATEGORY_LIST_CACHE["expires_at"] = 0
    return ok


def merge_categories_to_tree(mappings: list) -> dict:
    """将旧分类名批量改为新路径：mappings = [ {"from": "早饭", "to": "正餐|早饭"}, ... ]。同时更新 category_aliases。"""
    if not mappings:
        return {"success": 0, "failed": 0, "errors": []}
    supabase = get_supabase_client()
    updated = 0
    failed = 0
    errors = []
    for m in mappings:
        from_name = (m.get("from") or "").strip()
        to_path = (m.get("to") or "").strip()
        if not from_name or not to_path:
            continue
        try:
            r = supabase.table("records").update({"category": to_path}).eq("category", from_name).execute()
            updated += len(r.data) if r.data else 0
            supabase.table("category_aliases").update({"category": to_path}).eq("category", from_name).execute()
        except Exception as e:
            failed += 1
            errors.append(f"{from_name}→{to_path}: {str(e)[:50]}")
    invalidate_records_cache()
    CATEGORY_ALIAS_CACHE["expires_at"] = 0
    CATEGORY_LIST_CACHE["expires_at"] = 0
    return {"success": updated, "failed": failed, "errors": errors}


def get_category_presets() -> list:
    """从设置中读取手动添加的类目预设（尚未有记录也可显示）"""
    raw = get_setting("category_presets", "").strip()
    if not raw:
        return []
    try:
        out = json.loads(raw)
        return [str(p).strip() for p in out if str(p).strip()] if isinstance(out, list) else []
    except Exception:
        return []


def add_category_preset(path: str) -> bool:
    """添加一个类目预设（网页端「新增类目」用）"""
    path = (path or "").strip()
    if not path:
        return False
    presets = get_category_presets()
    if path in presets:
        return True
    presets.append(path)
    presets = sorted(set(presets))
    ok = set_setting("category_presets", json.dumps(presets, ensure_ascii=False))
    if ok:
        CATEGORY_LIST_CACHE["expires_at"] = 0
    return ok


def remove_category_preset(path: str) -> bool:
    """移除类目预设（仅移除预设，不删记录）"""
    path = (path or "").strip()
    presets = [p for p in get_category_presets() if p != path]
    ok = set_setting("category_presets", json.dumps(presets, ensure_ascii=False))
    if ok:
        CATEGORY_LIST_CACHE["expires_at"] = 0
    return ok


def get_all_categories() -> list:
    """获取所有分类 = 记录中出现的 + 手动添加的预设"""
    now = int(time.time())
    if CATEGORY_LIST_CACHE["value"] and now < CATEGORY_LIST_CACHE["expires_at"]:
        return CATEGORY_LIST_CACHE["value"]
    try:
        supabase = get_supabase_client()
        result = supabase.table("records").select("category").execute()
        categories = set()
        for r in result.data:
            cat = r.get("category", "").strip()
            if cat:
                categories.add(cat)
        for p in get_category_presets():
            if p:
                categories.add(p)
        sorted_categories = sorted(list(categories))
        CATEGORY_LIST_CACHE["value"] = sorted_categories
        CATEGORY_LIST_CACHE["expires_at"] = now + CATEGORY_LIST_CACHE_TTL
        return sorted_categories
    except Exception as e:
        print(f"获取分类列表错误: {str(e)[:100]}")
        if CATEGORY_LIST_CACHE["value"]:
            return CATEGORY_LIST_CACHE["value"]
        return []


def get_category_stats() -> list:
    """获取分类统计（含记录数）"""
    try:
        supabase = get_supabase_client()
        result = supabase.table("records").select("category").execute()
        category_count = {}
        for r in result.data:
            cat = r.get("category", "").strip()
            if cat:
                category_count[cat] = category_count.get(cat, 0) + 1
        return [{"category": cat, "count": count} for cat, count in sorted(category_count.items())]
    except Exception as e:
        print(f"分类统计错误: {str(e)[:100]}")
        return []


def build_category_excel_bytes() -> bytes:
    """导出分类管理 Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "分类管理"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    center_align = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9")
    )

    ws.append(["使用说明"])
    ws.append(["1. 修改「新分类名」列，保持「当前分类名」不变"])
    ws.append(["2. 保存后上传到 /api/import_categories 批量重命名"])
    ws.append(["3. 系统会自动更新所有历史记录"])
    ws.append([])

    ws.append(["当前分类名", "记录数", "新分类名"])
    for cell in ws[ws.max_row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = border

    stats = get_category_stats()
    for item in stats:
        ws.append([item["category"], item["count"], item["category"]])

    ws.freeze_panes = "A7"
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 20

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=3):
        for cell in row:
            if cell.value is None:
                continue
            cell.border = border
            if cell.row <= 4:
                cell.font = Font(color="FF0000")

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return bio.read()


def parse_category_excel(file_bytes: bytes) -> dict:
    """解析分类管理 Excel"""
    try:
        wb = load_workbook(io.BytesIO(file_bytes))
        if "分类管理" not in wb.sheetnames:
            return {"error": "no_category_sheet"}
        
        ws = wb["分类管理"]
        rows = list(ws.iter_rows(min_row=7, values_only=True))
        renames = []
        for row in rows:
            if not row or len(row) < 3:
                continue
            old_name, count, new_name = row[:3]
            if not old_name or not new_name:
                continue
            old_name = str(old_name).strip()
            new_name = str(new_name).strip()
            if old_name != new_name:
                renames.append({"old_name": old_name, "new_name": new_name})
        return {"renames": renames}
    except Exception as e:
        print(f"解析分类 Excel 错误: {str(e)[:100]}")
        return {"error": "parse_failed"}


def build_category_mapping_excel_bytes() -> bytes:
    """导出「分类映射表」模板：原分类、新分类。支持一级/二级/三级，新分类用----分隔如 正餐----晚餐----外卖。"""
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "分类映射"
        ws.append(["使用说明"])
        ws.append(["1. 本表为当前所有类目。修改「新分类」列即可合并，可写一级(正餐)或二级(正餐----早餐)或三级(正餐----晚餐----外卖)"])
        ws.append(["2. 也可单列写「新分类----原分类」，如 正餐----早餐、出行----打车；新分类里再用----表示多级"])
        ws.append(["3. 若删除某一行，上传后该分类的记录会进入「待处理」，需在网页上为每组选择要归入的分类"])
        ws.append(["4. 不要删表头"])
        ws.append([])
        ws.append(["原分类", "新分类"])
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="4F81BD")
        for col in range(1, 3):
            c = ws.cell(row=ws.max_row, column=col)
            c.font = header_font
            c.fill = header_fill
        stats = get_category_stats()
        for item in stats:
            cat = item.get("category") or ""
            cat_str = str(cat).strip() if cat is not None else ""
            if cat_str:
                ws.append([cat_str, cat_str])
        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)
        return bio.read()
    except Exception as e:
        print(f"导出映射表 build 错误: {str(e)[:200]}")
        raise


def _normalize_new_category(s: str) -> str:
    """新分类里用----表示多级，存库用|。如 正餐----晚餐----外卖 -> 正餐|晚餐|外卖"""
    if not s or not isinstance(s, str):
        return (s or "").strip()
    return s.strip().replace("----", "|")


def parse_category_mapping_excel(file_bytes: bytes) -> dict:
    """解析分类映射表 Excel。支持两列「原分类、新分类」或单列「新分类----原分类」。新分类可多级如 正餐----晚餐----外卖。"""
    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        sheet_name = "分类映射" if "分类映射" in wb.sheetnames else (wb.sheetnames[0] if wb.sheetnames else "")
        if not sheet_name:
            return {"error": "no_sheet"}
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        mappings = []
        in_header = set()
        for row in rows:
            if not row:
                continue
            cell0 = str(row[0] or "").strip()
            if not cell0 or cell0 in ("原分类", "新分类", "新分类----原分类"):
                continue
            if "----" in cell0:
                parts = cell0.rsplit("----", 1)
                new_cat_raw = (parts[0] or "").strip()
                orig = (parts[1] or "").strip()
                if not orig:
                    orig = new_cat_raw
                if not new_cat_raw:
                    new_cat_raw = orig
                new_cat = _normalize_new_category(new_cat_raw)
                if not new_cat:
                    new_cat = orig
            else:
                orig = cell0
                new_cat_raw = str(row[1] or "").strip() if len(row) >= 2 else orig
                if not new_cat_raw:
                    new_cat_raw = orig
                new_cat = _normalize_new_category(new_cat_raw)
                if not new_cat:
                    new_cat = orig
            in_header.add(orig)
            mappings.append({"from": orig, "to": new_cat})
        return {"mappings": mappings, "origins_in_file": list(in_header)}
    except Exception as e:
        print(f"解析映射表错误: {str(e)[:100]}")
        return {"error": "parse_failed"}


def batch_rename_categories(renames: list) -> dict:
    """批量重命名分类"""
    success = 0
    failed = []
    for item in renames:
        result = rename_category(item["old_name"], item["new_name"])
        if result.get("success"):
            success += 1
        else:
            failed.append(item["old_name"])
    return {"success": success, "failed": failed}


def clear_category_aliases(category_name: str) -> bool:
    """清除某分类下的所有别名（该分类无记录时用于「删除分类」）"""
    try:
        supabase = get_supabase_client()
        supabase.table("category_aliases").delete().eq("category", category_name.strip()).execute()
        CATEGORY_ALIAS_CACHE["value"] = {}
        CATEGORY_ALIAS_CACHE["expires_at"] = 0
        return True
    except Exception as e:
        print(f"清除分类别名错误: {str(e)[:100]}")
        return False


def add_months(dt: datetime, months: int) -> datetime:
    """按月偏移"""
    year = dt.year + (dt.month - 1 + months) // 12
    month = (dt.month - 1 + months) % 12 + 1
    return dt.replace(year=year, month=month, day=1)


def build_dashboard_text() -> str:
    """统计面板：月/周/年趋势 + 分类占比"""
    now = datetime.now(LOCAL_TZ)
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    stats_month = get_statistics(start_date=month_start, end_date=now)

    lines = ["📊 统计面板", ""]

    # 分类占比（本月）
    if stats_month["count"] > 0:
        lines.append("🧭 本月分类占比")
        total = stats_month["total"] or 0
        sorted_cats = sorted(stats_month["by_category"].items(), key=lambda x: -x[1])
        for cat, amount in sorted_cats:
            percent = (amount / total * 100) if total else 0
            lines.append(f"{cat} {amount:.2f}元 ({percent:.1f}%)")
        lines.append("")
    else:
        lines.append("🧭 本月分类占比：暂无数据")
        lines.append("")

    # 趋势数据（近4周 / 近6月 / 近3年）
    week_start = now.replace(hour=0, minute=0, second=0, microsecond=0) - timedelta(days=now.weekday())
    start_year = now.replace(year=now.year - 2, month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
    records = get_records(start_date=start_year - timedelta(days=1), end_date=now + timedelta(days=1))
    records = filter_records_by_local_range(records, start_year, now + timedelta(days=1))

    weekly_totals = {}
    monthly_totals = {}
    yearly_totals = {}

    for r in records:
        dt = to_local_datetime(r["created_at"])
        amount = float(r["amount"])
        week_key = (dt - timedelta(days=dt.weekday())).strftime("%Y-%m-%d")
        month_key = dt.strftime("%Y-%m")
        year_key = str(dt.year)
        weekly_totals[week_key] = weekly_totals.get(week_key, 0) + amount
        monthly_totals[month_key] = monthly_totals.get(month_key, 0) + amount
        yearly_totals[year_key] = yearly_totals.get(year_key, 0) + amount

    # 周趋势（近4周）
    lines.append("📅 近4周趋势")
    for i in range(3, -1, -1):
        ws = week_start - timedelta(weeks=i)
        we = ws + timedelta(days=6)
        key = ws.strftime("%Y-%m-%d")
        total = weekly_totals.get(key, 0)
        lines.append(f"{ws.strftime('%m/%d')}-{we.strftime('%m/%d')} {total:.2f}元")
    lines.append("")

    # 月趋势（近6月）
    lines.append("🗓️ 近6月趋势")
    first_month = add_months(month_start, -5)
    for i in range(6):
        current = add_months(first_month, i)
        key = current.strftime("%Y-%m")
        total = monthly_totals.get(key, 0)
        lines.append(f"{current.strftime('%Y-%m')} {total:.2f}元")
    lines.append("")

    # 年趋势（近3年）
    lines.append("📈 近3年趋势")
    for year in range(now.year - 2, now.year + 1):
        total = yearly_totals.get(str(year), 0)
        lines.append(f"{year}年 {total:.2f}元")

    return "\n".join(lines)


ACCESS_TOKEN_CACHE = {"value": "", "expires_at": 0}


def get_access_token() -> str:
    """获取公众号 access_token（缓存）"""
    now = int(time.time())
    if ACCESS_TOKEN_CACHE["value"] and now < ACCESS_TOKEN_CACHE["expires_at"]:
        return ACCESS_TOKEN_CACHE["value"]

    if not APPID or not APPSECRET:
        raise RuntimeError("missing app credentials")

    url = "https://api.weixin.qq.com/cgi-bin/token"
    params = {"grant_type": "client_credential", "appid": APPID, "secret": APPSECRET}
    response = httpx.get(url, params=params, timeout=10.0)
    response.raise_for_status()
    data = response.json()
    token = data.get("access_token", "")
    expires_in = int(data.get("expires_in", 0))
    if not token:
        raise RuntimeError("access_token missing")
    ACCESS_TOKEN_CACHE["value"] = token
    ACCESS_TOKEN_CACHE["expires_at"] = now + max(0, expires_in - 120)
    return token


def send_text_message(openid: str, text: str) -> bool:
    """客服消息推送（用户48小时内互动有效）"""
    token = get_access_token()
    url = f"https://api.weixin.qq.com/cgi-bin/message/custom/send?access_token={token}"
    payload = {
        "touser": openid,
        "msgtype": "text",
        "text": {"content": text}
    }
    response = httpx.post(url, json=payload, timeout=10.0)
    response.raise_for_status()
    data = response.json()
    return data.get("errcode") == 0


def format_debts(debts: list) -> str:
    """格式化外债列表"""
    if not debts:
        return "📌 外债总览：暂无欠款"

    total = sum(float(d.get("amount", 0)) for d in debts)
    lines = ["📌 外债总览（我欠别人）"]
    for d in debts:
        lines.append(f"  • {d['name']}：{float(d['amount']):.2f} 元")
    lines.append(f"合计：{total:.2f} 元")
    return "\n".join(lines)


def get_help_text() -> str:
    """返回帮助信息（支持自定义）"""
    # 尝试获取自定义帮助文本
    custom_help = get_setting("custom_help_text", "")
    if custom_help.strip():
        return custom_help
    
    # 默认帮助文本
    return """📖 记账机器人使用指南

【记账】
发送：描述 金额（AI自动分类）
例如：早餐 15 / 打车 30 / 咖啡18

【批量补记】
发送多行（每行可带日期）：
补记
5/5 早餐 8
5/5 午餐 25
5/6 咖啡 15

【查询统计】
今日 / 昨日 / 七天 / 本周 / 本月
统计 1月 / 统计面板

【明细】明细 / 明细 昨天

【修改/删除】
改 1 描述 金额 / 删 2 / 删 1-4

【补记单条】补记 昨天 买烟 50

【导出】导出 本月 / 导出 全部

【管理后台】发送：网页
（支持快速补录、导入账单CSV）

💡 发送 帮助 查看完整指南"""


# ============ 处理消息 ============
def handle_message(openid: str, nickname: str, content: str) -> str:
    """处理用户消息，返回回复内容"""
    content = content.strip()

    # 分类选择处理
    pending_pick = PENDING_CATEGORY_PICKS.get(openid)
    if pending_pick:
        if time.time() - pending_pick["ts"] > PENDING_CATEGORY_TTL:
            PENDING_CATEGORY_PICKS.pop(openid, None)
        elif content in ["取消", "取消分类"]:
            PENDING_CATEGORY_PICKS.pop(openid, None)
            return "✅ 已取消分类选择"
        elif content.isdigit():
            idx = int(content)
            categories = pending_pick["categories"]
            if 1 <= idx <= len(categories):
                category = categories[idx - 1]
                add_record(
                    openid=openid,
                    nickname=nickname,
                    amount=pending_pick["amount"],
                    category=category,
                    description=pending_pick["description"]
                )
                add_category_alias(pending_pick["description"], category)
                PENDING_CATEGORY_PICKS.pop(openid, None)
                return (
                    f"✅ 记账成功！\n{pending_pick['description']}：{pending_pick['amount']:.2f} 元\n"
                    f"分类：{category}\n已记住，下次将自动归类"
                )
            return build_category_pick_prompt(pending_pick["description"], pending_pick["amount"], categories)
        else:
            return build_category_pick_prompt(pending_pick["description"], pending_pick["amount"], pending_pick["categories"])

    # 批量补记 / 批量记账
    # 支持格式1：补记\n5/5 早餐 8\n5/5 午餐 25
    # 支持格式2：批量\n早餐 8;午餐 25;打车 30
    # 支持格式3（日期分组）：
    #   5.1
    #   1688先采后付	98.26
    #   保险	12.5
    #   5.2
    #   小白修车	45
    is_batch_backfill = content.startswith("补记") and ("\n" in content or "；" in content or ";" in content)
    is_batch = "批量" in content or "\n" in content or "；" in content or ";" in content

    if is_batch_backfill or is_batch:
        raw = content
        if is_batch_backfill:
            raw = content[len("补记"):].strip()
        else:
            raw = content.replace("批量", "").strip()
        raw = raw.replace("；", "\n").replace(";", "\n")
        lines = [l.strip() for l in raw.splitlines() if l.strip()]
        if len(lines) >= 2:
            success = 0
            failed = []
            current_date = None

            def _try_parse_date_only(line_text: str):
                """尝试解析纯日期行，如 5.1、5/1、5-1、05.01、5月1日、2026.5.1"""
                m = re.match(r'^(?:(\d{4})[./\-年])?\s*(\d{1,2})[./\-月]\s*(\d{1,2})[日号]?\s*$', line_text)
                if m:
                    year_str, month_str, day_str = m.groups()
                    now = datetime.now(LOCAL_TZ)
                    year = int(year_str) if year_str else now.year
                    month = int(month_str)
                    day = int(day_str)
                    dt = datetime(year, month, day, 12, 0, 0, tzinfo=LOCAL_TZ)
                    if dt > now:
                        dt = dt.replace(year=year - 1)
                    return dt
                return None

            def _try_parse_date_prefix(line_text: str):
                """尝试解析行首日期+内容，如 5/5 早餐 8"""
                m = re.match(
                    r'^(?:(\d{4})[./\-年])?\s*(\d{1,2})[./\-月]\s*(\d{1,2})[日号]?\s+(.+)$',
                    line_text
                )
                if m:
                    year_str, month_str, day_str, rest = m.groups()
                    now = datetime.now(LOCAL_TZ)
                    year = int(year_str) if year_str else now.year
                    month = int(month_str)
                    day = int(day_str)
                    dt = datetime(year, month, day, 12, 0, 0, tzinfo=LOCAL_TZ)
                    if dt > now:
                        dt = dt.replace(year=year - 1)
                    return dt, rest.strip()
                return None, None

            def _parse_tab_record(line_text: str) -> dict:
                """解析 tab 分隔的记录行，如 '1688先采后付\t98.26'"""
                if "\t" in line_text:
                    parts = line_text.split("\t")
                    if len(parts) >= 2:
                        desc = parts[0].strip()
                        amount_str = parts[-1].strip()
                        try:
                            amount = float(amount_str)
                            if desc and amount > 0:
                                return {"type": "record", "amount": amount, "description": desc, "category": desc, "explicit_category": False}
                        except ValueError:
                            pass
                return {"type": "unknown"}

            parsed_items = []
            for line in lines:
                record_date = None
                record_text = line

                date_only = _try_parse_date_only(line)
                if date_only:
                    current_date = date_only
                    continue

                prefix_date, prefix_rest = _try_parse_date_prefix(line)
                if prefix_date:
                    record_date = prefix_date
                    current_date = prefix_date
                    record_text = prefix_rest

                if record_date is None and current_date is not None:
                    record_date = current_date

                parsed_line = _parse_tab_record(record_text)
                if parsed_line["type"] != "record":
                    parsed_line = parse_record_text(record_text)

                if parsed_line["type"] == "record":
                    parsed_items.append({"parsed": parsed_line, "date": record_date, "line": line})
                else:
                    failed.append(line)

            if not parsed_items:
                pass
            else:
                descs_need_ai = []
                for item in parsed_items:
                    desc = item["parsed"]["description"]
                    alias = match_alias_category(desc)
                    item["category"] = alias if alias else None
                    if not alias:
                        descs_need_ai.append(desc)

                ai_categories = {}
                if descs_need_ai:
                    ai_categories = ai_batch_classify(descs_need_ai)

                for item in parsed_items:
                    try:
                        desc = item["parsed"]["description"]
                        category = item["category"] or ai_categories.get(desc, "") or "其他"
                        _ensure_category_in_tree(category, "")
                        add_category_alias(desc, category)
                        add_record(
                            openid=openid,
                            nickname=nickname,
                            amount=item["parsed"]["amount"],
                            category=category,
                            description=desc,
                            created_at=item["date"]
                        )
                        success += 1
                    except Exception:
                        failed.append(item["line"])

                msg = f"✅ 批量记账完成：成功{success}条"
                if failed:
                    msg += f"\n❌ 失败{len(failed)}条：\n" + "\n".join(failed[:5])
                return msg

    # 待确认删除的状态处理
    pending_del = PENDING_DELETES.get(openid)
    if pending_del:
        if time.time() - pending_del["ts"] > PENDING_DELETE_TTL:
            PENDING_DELETES.pop(openid, None)
        elif content in ["确认删", "确认删除", "确认"]:
            try:
                deleted = 0
                for record in pending_del["items"]:
                    archive_deleted_record(record, deleted_by=openid)
                    if delete_record(record["id"]):
                        deleted += 1
                PENDING_DELETES.pop(openid, None)
                if deleted == 0:
                    return "删除失败，数据库可能有权限问题"
                return f"✅ 已删除 {deleted} 条记录"
            except Exception as e:
                print(f"确认删除失败: {str(e)[:100]}")
                return "删除失败，请稍后重试"
        elif content in ["取消删", "取消删除", "取消"]:
            PENDING_DELETES.pop(openid, None)
            return "✅ 已取消删除"

    # 所有消息交给 AI 理解和处理
    return handle_ai_intent(openid, nickname, content)


# ============ 微信公众号验证 ============
def check_signature(signature, timestamp, nonce):
    """验证微信服务器签名"""
    tmp_arr = [TOKEN, timestamp, nonce]
    tmp_arr.sort()
    tmp_str = ''.join(tmp_arr)
    tmp_str = hashlib.sha1(tmp_str.encode('utf-8')).hexdigest()
    return tmp_str == signature


# ============ API 路由 ============
@app.get("/api/wechat")
async def verify(request: Request):
    """微信公众号 URL 验证"""
    try:
        params = dict(request.query_params)
        signature = params.get("signature", "")
        timestamp = params.get("timestamp", "")
        nonce = params.get("nonce", "")
        echostr = params.get("echostr", "")
        
        if check_signature(signature, timestamp, nonce):
            return Response(content=echostr, media_type="text/plain")
        else:
            return Response(content="verify failed", status_code=403)
    except Exception as e:
        print(f"验证错误: {str(e)[:100]}")
        return Response(content="error", status_code=500)


@app.get("/api/health")
async def health():
    """健康检查（保活用）"""
    return Response(content="ok", media_type="text/plain")


@app.get("/api/debug_db")
async def debug_db():
    """诊断数据库连接"""
    import traceback
    info = {
        "supabase_url": SUPABASE_URL[:30] + "..." if SUPABASE_URL else "(empty)",
        "supabase_key_prefix": SUPABASE_KEY[:20] + "..." if SUPABASE_KEY else "(empty)",
        "deepseek_key": "set" if DEEPSEEK_API_KEY else "(empty)",
    }
    try:
        supabase = get_supabase_client()
        result = supabase.table("records").select("id").limit(1).execute()
        info["db_status"] = "OK"
        info["db_result"] = str(result.data)
    except Exception as e:
        info["db_status"] = "ERROR"
        info["db_error"] = str(e)
        info["db_traceback"] = traceback.format_exc()[-500:]
    return info


@app.head("/api/health")
async def health_head():
    """健康检查（HEAD）"""
    return Response(content="", status_code=200)


@app.post("/api/wechat")
async def webhook(request: Request):
    """接收微信公众号消息"""
    try:
        body = await request.body()
        body_str = body.decode("utf-8")
        
        # 解析 XML
        from xml.etree import ElementTree as ET
        xml_tree = ET.fromstring(body_str)
        
        msg_type = xml_tree.find("MsgType").text
        from_user = xml_tree.find("FromUserName").text
        msg_id_node = xml_tree.find("MsgId")
        msg_id = msg_id_node.text if msg_id_node is not None else ""
        
        if msg_type not in ("text", "image"):
            return Response(content="success", media_type="text/plain")

        if msg_id and is_duplicate_message(msg_id):
            return Response(content="success", media_type="text/plain")
        
        nickname = from_user[:8]
        
        if msg_type == "image":
            pic_url_node = xml_tree.find("PicUrl")
            pic_url = pic_url_node.text if pic_url_node is not None else ""
            if not pic_url:
                reply_content = "❌ 未能获取图片地址"
            else:
                reply_content = handle_image_message(from_user, nickname, pic_url)
        else:
            content = xml_tree.find("Content").text
            reply_content = handle_message(from_user, nickname, content)

        if msg_id:
            record_message_id(msg_id)
        
        # 构造回复 XML
        to_user = xml_tree.find("FromUserName").text
        from_user_name = xml_tree.find("ToUserName").text
        create_time = int(time.time())
        
        reply_xml = f"""<xml>
<ToUserName><![CDATA[{to_user}]]></ToUserName>
<FromUserName><![CDATA[{from_user_name}]]></FromUserName>
<CreateTime>{create_time}</CreateTime>
<MsgType><![CDATA[text]]></MsgType>
<Content><![CDATA[{reply_content}]]></Content>
</xml>"""
        
        return Response(content=reply_xml, media_type="application/xml")
    except Exception as e:
        print(f"处理消息错误: {str(e)[:100]}")
        return Response(content="success", media_type="text/plain")


@app.get("/api/export")
async def export_excel(request: Request):
    """导出 Excel"""
    try:
        params = dict(request.query_params)
        openid = params.get("openid", "")
        period = params.get("period", "")
        ts = params.get("ts", "")
        sig = params.get("sig", "")

        if not verify_export_signature(openid, period, ts, sig):
            return Response(content="签名验证失败", status_code=403)

        if period.startswith("month:"):
            month_text = period.split("month:", 1)[1]
            month_range = parse_month_token(month_text)
            if not month_range:
                return Response(content="月份格式错误", status_code=400)
            start_date, end_date, _ = month_range
        else:
            start_date, end_date = get_date_range(period)
        
        if not start_date or not end_date:
            return Response(content="日期范围错误", status_code=400)
            
        records = get_records(start_date=start_date - timedelta(days=1), end_date=end_date + timedelta(days=1))
        records = filter_records_by_local_range(records, start_date, end_date)
        # 全部导出时增加限制到10000条
        limit = 10000 if period == "all" else 1000
        data = build_export_excel_bytes(records, start_date, end_date, limit=limit)
        
        # 生成文件名：导出时间_范围.xlsx（使用英文避免编码问题）
        export_time = datetime.now(LOCAL_TZ).strftime("%Y%m%d_%H%M%S")
        if period.startswith("month:"):
            month_text = period.split("month:", 1)[1]
            range_text = month_text.replace("-", "")
        elif period == "all":
            range_text = "all"
        else:
            range_text = period
        filename = f"records_{export_time}_{range_text}.xlsx"
        headers = {
            "Content-Disposition": f'attachment; filename="{filename}"'
        }
        return StreamingResponse(
            io.BytesIO(data),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers
        )
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        print(f"导出错误详情: {error_detail}")
        return Response(content=f"导出失败: {str(e)}", status_code=500)


@app.get("/api/import", response_class=HTMLResponse)
async def import_page():
    """上传页面"""
    import os
    upload_html_path = os.path.join(os.path.dirname(__file__), "upload.html")
    try:
        with open(upload_html_path, "r", encoding="utf-8") as f:
            return f.read()
    except FileNotFoundError:
        return "<h1>上传页面未找到</h1>"


@app.get("/api/import_categories_page", response_class=HTMLResponse)
async def import_categories_page():
    """批量修改分类页面"""
    import os
    upload_html_path = os.path.join(os.path.dirname(__file__), "upload_categories.html")
    try:
        with open(upload_html_path, "r", encoding="utf-8") as f:
            return f.read()
    except FileNotFoundError:
        return "<h1>上传页面未找到</h1>"


@app.post("/api/import")
async def import_excel(file: UploadFile = File(...), request: Request = None):
    """批量导入修改后的 Excel（从明细表读取）"""
    try:
        file_bytes = await file.read()
        parsed = parse_import_excel(file_bytes)
        if parsed.get("error"):
            return Response(content=parsed["error"], status_code=400)
        
        updates = parsed.get("updates", [])
        if not updates:
            return Response(content="no_valid_records", status_code=400)
        
        result = batch_update_records(updates)
        return Response(
            content=f"ok: {result['success']} success, {len(result['failed'])} failed",
            media_type="text/plain"
        )
    except Exception as e:
        print(f"导入错误: {str(e)[:100]}")
        return Response(content="error", status_code=500)


@app.post("/api/admin/import")
async def admin_import_excel(file: UploadFile = File(...), payload: dict = Depends(verify_admin_token)):
    """管理后台：批量导入修改后的 Excel（按 ID 更新已有记录）"""
    try:
        file_bytes = await file.read()
        parsed = parse_import_excel(file_bytes)
        if parsed.get("error"):
            return {"success": False, "error": parsed["error"]}
        updates = parsed.get("updates", [])
        if not updates:
            return {"success": False, "error": "没有可更新的记录"}
        result = batch_update_records(updates)
        return {"success": True, "updated": result["success"], "failed": len(result["failed"]), "message": f"已更新 {result['success']} 条，失败 {len(result['failed'])} 条"}
    except Exception as e:
        print(f"导入错误: {str(e)[:100]}")
        return {"success": False, "error": str(e)}


def parse_wechat_csv(file_bytes: bytes) -> list:
    """解析微信支付账单 CSV"""
    import csv
    records = []
    try:
        text = file_bytes.decode("utf-8")
    except UnicodeDecodeError:
        text = file_bytes.decode("gbk", errors="ignore")
    lines = text.splitlines()
    header_idx = -1
    for i, line in enumerate(lines):
        if "交易时间" in line and "交易类型" in line:
            header_idx = i
            break
    if header_idx < 0:
        return []
    reader = csv.reader(lines[header_idx:])
    headers = next(reader, [])
    headers = [h.strip().strip("\ufeff") for h in headers]
    time_col = next((i for i, h in enumerate(headers) if "交易时间" in h), None)
    type_col = next((i for i, h in enumerate(headers) if "交易类型" in h), None)
    counterpart_col = next((i for i, h in enumerate(headers) if "交易对方" in h), None)
    goods_col = next((i for i, h in enumerate(headers) if "商品" in h), None)
    amount_col = next((i for i, h in enumerate(headers) if "金额" in h), None)
    inout_col = next((i for i, h in enumerate(headers) if "收/支" in h or "收／支" in h), None)

    for row in reader:
        if not row or len(row) < max(filter(None, [time_col, amount_col, goods_col]), default=0) + 1:
            continue
        try:
            inout = row[inout_col].strip() if inout_col is not None else ""
            if "支出" not in inout:
                continue
            amount_str = row[amount_col].strip().replace("¥", "").replace(",", "").strip() if amount_col is not None else ""
            amount = float(amount_str)
            if amount <= 0:
                continue
            time_str = row[time_col].strip() if time_col is not None else ""
            description = row[goods_col].strip() if goods_col is not None else ""
            counterpart = row[counterpart_col].strip() if counterpart_col is not None else ""
            if not description or description == "/":
                description = counterpart or "未知"
            created_at = None
            if time_str:
                for fmt in ["%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M"]:
                    try:
                        created_at = datetime.strptime(time_str, fmt).replace(tzinfo=LOCAL_TZ)
                        break
                    except ValueError:
                        continue
            records.append({
                "amount": amount,
                "description": description,
                "counterpart": counterpart,
                "created_at": created_at
            })
        except (ValueError, IndexError):
            continue
    return records


def parse_alipay_csv(file_bytes: bytes) -> list:
    """解析支付宝账单 CSV"""
    import csv
    records = []
    try:
        text = file_bytes.decode("utf-8")
    except UnicodeDecodeError:
        text = file_bytes.decode("gbk", errors="ignore")
    lines = text.splitlines()
    header_idx = -1
    for i, line in enumerate(lines):
        if "交易时间" in line and ("商品说明" in line or "商品名称" in line):
            header_idx = i
            break
        if "交易创建时间" in line:
            header_idx = i
            break
    if header_idx < 0:
        return []
    reader = csv.reader(lines[header_idx:])
    headers = next(reader, [])
    headers = [h.strip().strip("\ufeff").strip("\t") for h in headers]
    time_col = next((i for i, h in enumerate(headers) if "交易时间" in h or "交易创建时间" in h), None)
    goods_col = next((i for i, h in enumerate(headers) if "商品" in h), None)
    counterpart_col = next((i for i, h in enumerate(headers) if "交易对方" in h or "对方" in h), None)
    amount_col = next((i for i, h in enumerate(headers) if "金额" in h), None)
    inout_col = next((i for i, h in enumerate(headers) if "收/支" in h or "收／支" in h), None)

    for row in reader:
        if not row or len(row) < 4:
            continue
        row = [c.strip().strip("\t") for c in row]
        try:
            inout = row[inout_col].strip() if inout_col is not None else ""
            if "支出" not in inout:
                continue
            amount_str = row[amount_col].strip().replace(",", "") if amount_col is not None else ""
            amount = float(amount_str)
            if amount <= 0:
                continue
            time_str = row[time_col].strip() if time_col is not None else ""
            description = row[goods_col].strip() if goods_col is not None else ""
            counterpart = row[counterpart_col].strip() if counterpart_col is not None else ""
            if not description or description == "/":
                description = counterpart or "未知"
            created_at = None
            if time_str:
                for fmt in ["%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d %H:%M"]:
                    try:
                        created_at = datetime.strptime(time_str, fmt).replace(tzinfo=LOCAL_TZ)
                        break
                    except ValueError:
                        continue
            records.append({
                "amount": amount,
                "description": description,
                "counterpart": counterpart,
                "created_at": created_at
            })
        except (ValueError, IndexError):
            continue
    return records


@app.get("/api/admin/quick_entry", response_class=HTMLResponse)
async def admin_quick_entry_page():
    """快速补录页面"""
    import os
    html_path = os.path.join(os.path.dirname(__file__), "quick_entry.html")
    try:
        with open(html_path, "r", encoding="utf-8") as f:
            return f.read()
    except FileNotFoundError:
        return "<h1>快速补录页面未找到</h1>"


@app.post("/api/admin/quick_entry")
async def admin_quick_entry(request: Request, payload: dict = Depends(verify_admin_token)):
    """快速补录：批量新增记录"""
    try:
        data = await request.json()
        items = data.get("items", [])
        if not items:
            return {"success": False, "error": "没有记录"}
        categories_for_ai = get_category_candidates()
        success = 0
        failed = 0
        for item in items:
            try:
                description = (item.get("description") or "").strip()
                amount = float(item.get("amount", 0))
                category = (item.get("category") or "").strip()
                date_str = (item.get("date") or "").strip()
                time_str = (item.get("time") or "12:00").strip()
                if not description or amount <= 0:
                    failed += 1
                    continue
                if not category:
                    category = match_alias_category(description)
                if not category:
                    category = ai_classify(description, categories_for_ai)
                if not category:
                    category = "其他"
                created_at = None
                if date_str:
                    try:
                        dt_str = f"{date_str} {time_str}"
                        created_at = datetime.strptime(dt_str, "%Y-%m-%d %H:%M").replace(tzinfo=LOCAL_TZ)
                    except ValueError:
                        pass
                add_record(
                    openid="admin",
                    nickname="管理员",
                    amount=amount,
                    category=category,
                    description=description,
                    created_at=created_at
                )
                add_category_alias(description, category)
                success += 1
            except Exception:
                failed += 1
        return {"success": True, "added": success, "failed": failed, "message": f"成功录入 {success} 条，失败 {failed} 条"}
    except Exception as e:
        print(f"快速补录错误: {str(e)[:100]}")
        return {"success": False, "error": str(e)}


@app.post("/api/admin/import_csv")
async def admin_import_csv(file: UploadFile = File(...), payload: dict = Depends(verify_admin_token)):
    """导入微信/支付宝账单 CSV"""
    try:
        file_bytes = await file.read()
        filename = file.filename or ""
        text_preview = ""
        try:
            text_preview = file_bytes[:2000].decode("utf-8")
        except UnicodeDecodeError:
            text_preview = file_bytes[:2000].decode("gbk", errors="ignore")

        if "支付宝" in filename or "支付宝" in text_preview or "交易号" in text_preview:
            records = parse_alipay_csv(file_bytes)
        else:
            records = parse_wechat_csv(file_bytes)

        if not records:
            return {"success": False, "error": "未识别到有效的支出记录，请确认是微信或支付宝账单CSV", "records": []}

        categories_for_ai = get_category_candidates()
        for r in records:
            alias_cat = match_alias_category(r["description"])
            if alias_cat:
                r["category"] = alias_cat
            else:
                ai_cat = ai_classify(r["description"], categories_for_ai)
                r["category"] = ai_cat if ai_cat else ""
            r["date"] = r["created_at"].strftime("%Y-%m-%d") if r.get("created_at") else ""
            r["time"] = r["created_at"].strftime("%H:%M") if r.get("created_at") else ""

        preview = []
        for r in records[:200]:
            preview.append({
                "description": r["description"],
                "amount": r["amount"],
                "category": r["category"],
                "date": r.get("date", ""),
                "time": r.get("time", ""),
                "counterpart": r.get("counterpart", "")
            })
        return {"success": True, "records": preview, "total": len(records), "message": f"识别到 {len(records)} 条支出记录"}
    except Exception as e:
        print(f"CSV导入错误: {str(e)[:200]}")
        return {"success": False, "error": str(e), "records": []}


@app.post("/api/admin/confirm_csv_import")
async def admin_confirm_csv_import(request: Request, payload: dict = Depends(verify_admin_token)):
    """确认导入 CSV 账单记录"""
    try:
        data = await request.json()
        items = data.get("items", [])
        if not items:
            return {"success": False, "error": "没有记录"}
        success = 0
        failed = 0
        for item in items:
            try:
                description = (item.get("description") or "").strip()
                amount = float(item.get("amount", 0))
                category = (item.get("category") or "其他").strip()
                date_str = (item.get("date") or "").strip()
                time_str = (item.get("time") or "12:00").strip()
                if not description or amount <= 0:
                    failed += 1
                    continue
                created_at = None
                if date_str:
                    try:
                        dt_str = f"{date_str} {time_str}"
                        created_at = datetime.strptime(dt_str, "%Y-%m-%d %H:%M").replace(tzinfo=LOCAL_TZ)
                    except ValueError:
                        pass
                add_record(
                    openid="csv_import",
                    nickname="账单导入",
                    amount=amount,
                    category=category,
                    description=description,
                    created_at=created_at
                )
                add_category_alias(description, category)
                success += 1
            except Exception:
                failed += 1
        return {"success": True, "added": success, "failed": failed, "message": f"成功导入 {success} 条，失败 {failed} 条"}
    except Exception as e:
        print(f"确认CSV导入错误: {str(e)[:100]}")
        return {"success": False, "error": str(e)}


@app.get("/api/export_categories")
async def export_categories():
    """导出分类管理 Excel"""
    try:
        data = build_category_excel_bytes()
        filename = "categories.xlsx"
        from urllib.parse import quote
        encoded_filename = quote(filename)
        headers = {
            "Content-Disposition": f'attachment; filename="{filename}"; filename*=UTF-8\'\'{encoded_filename}',
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        }
        return StreamingResponse(io.BytesIO(data),
                                 media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 headers=headers)
    except Exception as e:
        print(f"导出分类错误: {str(e)[:100]}")
        return Response(content="error", status_code=500)


@app.post("/api/import_categories")
async def import_categories(file: UploadFile = File(...)):
    """批量导入修改后的分类 Excel"""
    try:
        file_bytes = await file.read()
        parsed = parse_category_excel(file_bytes)
        if parsed.get("error"):
            return Response(content=parsed["error"], status_code=400)
        
        renames = parsed.get("renames", [])
        if not renames:
            return Response(content="no_changes", status_code=400)
        
        result = batch_rename_categories(renames)
        return Response(
            content=f"ok: {result['success']} categories renamed, {len(result['failed'])} failed",
            media_type="text/plain"
        )
    except Exception as e:
        print(f"导入分类错误: {str(e)[:100]}")
        return Response(content="error", status_code=500)


@app.get("/api/report/weekly")
async def report_weekly(request: Request):
    """周报推送（需 REPORT_TOKEN）"""
    token = dict(request.query_params).get("token", "")
    if not REPORT_TOKEN or token != REPORT_TOKEN:
        return Response(content="invalid", status_code=403)
    try:
        text = build_report_text("7days", "近七天")
        subscribers = list_subscribers("weekly")
        success = 0
        for openid in subscribers:
            if send_text_message(openid, text):
                success += 1
        return Response(content=f"ok {success}/{len(subscribers)}", media_type="text/plain")
    except Exception as e:
        print(f"周报推送错误: {str(e)[:100]}")
        return Response(content="error", status_code=500)


@app.get("/api/report/monthly")
async def report_monthly(request: Request):
    """月报推送（需 REPORT_TOKEN）"""
    token = dict(request.query_params).get("token", "")
    if not REPORT_TOKEN or token != REPORT_TOKEN:
        return Response(content="invalid", status_code=403)
    try:
        text = build_report_text("30days", "近一个月")
        subscribers = list_subscribers("monthly")
        success = 0
        for openid in subscribers:
            if send_text_message(openid, text):
                success += 1
        return Response(content=f"ok {success}/{len(subscribers)}", media_type="text/plain")
    except Exception as e:
        print(f"月报推送错误: {str(e)[:100]}")
        return Response(content="error", status_code=500)


# ============ 管理后台 ============
@app.get("/api/admin", response_class=HTMLResponse)
async def admin_page():
    """管理后台页面"""
    import os
    admin_html_path = os.path.join(os.path.dirname(__file__), "admin.html")
    try:
        with open(admin_html_path, "r", encoding="utf-8") as f:
            return f.read()
    except FileNotFoundError:
        return "<h1>管理后台页面未找到</h1>"


@app.post("/api/admin/login")
async def admin_login(request: Request):
    """管理员登录（带失败次数限制）"""
    try:
        # 获取客户端 IP
        client_ip = request.client.host if request.client else "unknown"
        
        # 检查是否被锁定
        if client_ip in LOGIN_ATTEMPTS:
            attempts = LOGIN_ATTEMPTS[client_ip]
            if attempts["count"] >= MAX_LOGIN_ATTEMPTS:
                lockout_until = attempts.get("lockout_until", 0)
                if int(time.time()) < lockout_until:
                    remaining = lockout_until - int(time.time())
                    return {
                        "success": False,
                        "error": f"登录失败次数过多，请 {remaining} 秒后再试"
                    }
                else:
                    # 锁定时间已过，重置
                    LOGIN_ATTEMPTS.pop(client_ip, None)
        
        data = await request.json()
        password = data.get("password", "")
        
        if not ADMIN_PASSWORD:
            return {"success": False, "error": "未配置管理员密码"}
        
        if password != ADMIN_PASSWORD:
            # 记录失败次数
            if client_ip not in LOGIN_ATTEMPTS:
                LOGIN_ATTEMPTS[client_ip] = {"count": 0}
            LOGIN_ATTEMPTS[client_ip]["count"] += 1
            
            # 如果超过最大次数，锁定
            if LOGIN_ATTEMPTS[client_ip]["count"] >= MAX_LOGIN_ATTEMPTS:
                LOGIN_ATTEMPTS[client_ip]["lockout_until"] = int(time.time()) + LOGIN_LOCKOUT_TIME
                return {
                    "success": False,
                    "error": f"登录失败次数过多，账户已锁定 {LOGIN_LOCKOUT_TIME} 秒"
                }
            
            remaining = MAX_LOGIN_ATTEMPTS - LOGIN_ATTEMPTS[client_ip]["count"]
            return {
                "success": False,
                "error": f"密码错误，还可尝试 {remaining} 次"
            }
        
        # 登录成功，清除失败记录
        LOGIN_ATTEMPTS.pop(client_ip, None)
        
        # 生成Token（24小时过期）
        token = jwt.encode(
            {
                "type": "admin",
                "timestamp": int(time.time()),
                "exp": int(time.time()) + ADMIN_TOKEN_EXPIRY
            },
            ADMIN_SECRET,
            algorithm="HS256"
        )
        
        return {"success": True, "token": token}
    except Exception as e:
        print(f"登录错误: {str(e)[:100]}")
        return {"success": False, "error": "登录失败"}


@app.get("/api/admin/overview")
async def admin_overview(payload: dict = Depends(verify_admin_token)):
    """数据概览"""
    try:
        now = datetime.now(LOCAL_TZ)
        today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
        month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        
        # 总记录数（使用缓存）
        all_records = get_records_cached()
        total_count = len(all_records)
        
        # 今日记录
        today_records = filter_records_by_local_range(all_records, today_start, now + timedelta(days=1))
        today_amount = sum(float(r["amount"]) for r in today_records)
        
        # 本月记录
        month_records = filter_records_by_local_range(all_records, month_start, now + timedelta(days=1))
        month_amount = sum(float(r["amount"]) for r in month_records)
        
        # 分类数量
        categories = set(r["category"] for r in all_records)
        category_count = len(categories)
        
        return {
            "success": True,
            "total_count": total_count,
            "today_amount": today_amount,
            "month_amount": month_amount,
            "category_count": category_count
        }
    except Exception as e:
        print(f"概览错误: {str(e)[:100]}")
        return {"success": False, "error": str(e)}


@app.get("/api/admin/records")
async def admin_records(
    request: Request,
    payload: dict = Depends(verify_admin_token)
):
    """记录列表（分页、查询）"""
    try:
        params = dict(request.query_params)
        page = int(params.get("page", 1))
        page_size = int(params.get("page_size", 50))
        search = params.get("search", "")
        date_from = params.get("date_from", "")
        date_to = params.get("date_to", "")
        amount_min = params.get("amount_min", "")
        amount_max = params.get("amount_max", "")
        categories = params.get("categories", "")
        
        # 获取所有记录
        start_date = None
        end_date = None
        if date_from:
            start_date = datetime.strptime(date_from, "%Y-%m-%d").replace(tzinfo=LOCAL_TZ)
        if date_to:
            end_date = datetime.strptime(date_to, "%Y-%m-%d").replace(tzinfo=LOCAL_TZ) + timedelta(days=1)
        
        records = get_records(start_date=start_date, end_date=end_date)
        
        # 搜索过滤
        if search:
            search_lower = search.lower()
            records = [
                r for r in records
                if search_lower in r.get("description", "").lower() or
                   search_lower in r.get("category", "").lower()
            ]
        
        # 金额范围过滤
        if amount_min:
            min_amount = float(amount_min)
            records = [r for r in records if float(r.get("amount", 0)) >= min_amount]
        if amount_max:
            max_amount = float(amount_max)
            records = [r for r in records if float(r.get("amount", 0)) <= max_amount]
        
        # 分类过滤
        if categories:
            category_list = [c.strip() for c in categories.split(",") if c.strip()]
            records = [r for r in records if r.get("category", "") in category_list]
        
        # 格式化
        formatted = []
        for r in records:
            dt = to_local_datetime(r["created_at"])
            formatted.append({
                "id": r["id"],
                "date": dt.strftime("%Y-%m-%d"),
                "time": dt.strftime("%H:%M"),
                "description": r.get("description", ""),
                "amount": float(r.get("amount", 0)),
                "category": r.get("category", "")
            })
        
        # 分页
        total = len(formatted)
        start = (page - 1) * page_size
        end = start + page_size
        paginated = formatted[start:end]
        
        return {
            "success": True,
            "records": paginated,
            "total": total,
            "page": page,
            "page_size": page_size
        }
    except Exception as e:
        print(f"记录列表错误: {str(e)[:100]}")
        return {"success": False, "error": str(e)}


@app.put("/api/admin/records/{record_id}")
async def admin_update_record(
    record_id: int,
    request: Request,
    payload: dict = Depends(verify_admin_token)
):
    """更新记录"""
    try:
        data = await request.json()
        date_str = data.get("date", "")
        time_str = data.get("time", "")
        description = data.get("description", "")
        amount = float(data.get("amount", 0))
        category = data.get("category", "")
        
        # 合并日期时间
        if date_str and time_str:
            dt_str = f"{date_str} {time_str}"
            dt = datetime.strptime(dt_str, "%Y-%m-%d %H:%M")
            created_at = dt.replace(tzinfo=LOCAL_TZ)
        else:
            created_at = None
        
        update_data = {
            "description": description,
            "amount": amount,
            "category": category
        }
        if created_at:
            update_data["created_at"] = to_utc_iso(created_at)
        
        supabase = get_supabase_client()
        result = supabase.table("records").update(update_data).eq("id", record_id).execute()
        invalidate_records_cache()
        if result.data:
            # 查询同备注的记录数量，供前端判断是否需要批量修改映射
            same_desc_count = 0
            if description:
                try:
                    cnt = supabase.table("records").select("id", count="exact").eq("description", description).execute()
                    same_desc_count = cnt.count if cnt.count is not None else len(cnt.data)
                except Exception:
                    same_desc_count = 1
            return {"success": True, "same_desc_count": same_desc_count, "description": description, "category": category}
        else:
            return {"success": False, "error": "更新失败"}
    except Exception as e:
        print(f"更新记录错误: {str(e)[:100]}")
        return {"success": False, "error": str(e)}


@app.post("/api/admin/records/update_desc_category")
async def admin_update_desc_category(request: Request, payload: dict = Depends(verify_admin_token)):
    """将某备注的所有记录的分类都改为新分类，并更新别名映射"""
    try:
        data = await request.json()
        description = (data.get("description") or "").strip()
        new_category = (data.get("category") or "").strip()
        if not description or not new_category:
            return {"success": False, "error": "缺少参数"}
        supabase = get_supabase_client()
        result = supabase.table("records").update({"category": new_category}).eq("description", description).execute()
        updated = len(result.data) if result.data else 0
        # 更新别名映射
        add_category_alias(description, new_category)
        invalidate_records_cache()
        return {"success": True, "updated": updated}
    except Exception as e:
        print(f"批量修改备注分类错误: {str(e)[:100]}")
        return {"success": False, "error": str(e)}


@app.post("/api/admin/records/update_alias_only")
async def admin_update_alias_only(request: Request, payload: dict = Depends(verify_admin_token)):
    """仅更新别名映射（不修改历史记录）"""
    try:
        data = await request.json()
        description = (data.get("description") or "").strip()
        new_category = (data.get("category") or "").strip()
        if not description or not new_category:
            return {"success": False, "error": "缺少参数"}
        add_category_alias(description, new_category)
        return {"success": True}
    except Exception as e:
        return {"success": False, "error": str(e)}


@app.delete("/api/admin/records/{record_id}")
async def admin_delete_record(
    record_id: int,
    payload: dict = Depends(verify_admin_token)
):
    """删除记录"""
    try:
        if delete_record(record_id):
            return {"success": True}
        else:
            return {"success": False, "error": "删除失败，可能被 RLS 策略阻止"}
    except Exception as e:
        print(f"删除记录错误: {str(e)[:100]}")
        return {"success": False, "error": str(e)}


@app.get("/api/admin/stats")
async def admin_stats(
    request: Request,
    payload: dict = Depends(verify_admin_token)
):
    """统计数据（用于图表）"""
    try:
        params = dict(request.query_params)
        year = params.get("year", "")
        month = params.get("month", "")
        date = params.get("date", "")
        week = params.get("week", "")
        
        all_records = get_records_cached()
        
        # 根据筛选条件过滤记录
        if date:
            # 单日
            date_obj = datetime.strptime(date, "%Y-%m-%d").replace(tzinfo=LOCAL_TZ)
            date_start = date_obj
            date_end = date_obj + timedelta(days=1)
            records = filter_records_by_local_range(all_records, date_start, date_end)
        elif year and month:
            # 单月
            year_int = int(year)
            month_int = int(month)
            month_start = datetime(year_int, month_int, 1, 0, 0, 0, tzinfo=LOCAL_TZ)
            if month_int == 12:
                month_end = datetime(year_int + 1, 1, 1, 0, 0, 0, tzinfo=LOCAL_TZ)
            else:
                month_end = datetime(year_int, month_int + 1, 1, 0, 0, 0, tzinfo=LOCAL_TZ)
            records = filter_records_by_local_range(all_records, month_start, month_end)
        elif year:
            # 全年
            year_int = int(year)
            year_start = datetime(year_int, 1, 1, 0, 0, 0, tzinfo=LOCAL_TZ)
            year_end = datetime(year_int + 1, 1, 1, 0, 0, 0, tzinfo=LOCAL_TZ)
            records = filter_records_by_local_range(all_records, year_start, year_end)
        elif week:
            # 本周
            now = datetime.now(LOCAL_TZ)
            days_since_monday = now.weekday()
            week_start = (now - timedelta(days=days_since_monday)).replace(hour=0, minute=0, second=0, microsecond=0)
            week_end = now + timedelta(days=1)
            records = filter_records_by_local_range(all_records, week_start, week_end)
        else:
            # 全部
            records = all_records
        
        # 分类统计
        category_amounts = {}
        for r in records:
            cat = r.get("category", "其他")
            amount = float(r.get("amount", 0))
            category_amounts[cat] = category_amounts.get(cat, 0) + amount
        
        category_labels = list(category_amounts.keys())
        category_amounts_list = [category_amounts[c] for c in category_labels]
        
        # 趋势统计
        if date:
            # 按小时统计
            hourly_amounts = {}
            for r in records:
                dt = to_local_datetime(r["created_at"])
                hour_key = dt.strftime("%H:00")
                amount = float(r.get("amount", 0))
                hourly_amounts[hour_key] = hourly_amounts.get(hour_key, 0) + amount
            sorted_hours = sorted(hourly_amounts.keys())
            trend_labels = sorted_hours
            trend_amounts = [hourly_amounts[h] for h in sorted_hours]
        elif year and month:
            # 按日统计
            daily_amounts = {}
            for r in records:
                dt = to_local_datetime(r["created_at"])
                day_key = dt.strftime("%m-%d")
                amount = float(r.get("amount", 0))
                daily_amounts[day_key] = daily_amounts.get(day_key, 0) + amount
            sorted_days = sorted(daily_amounts.keys())
            trend_labels = sorted_days
            trend_amounts = [daily_amounts[d] for d in sorted_days]
        elif week:
            # 按日统计（本周）
            daily_amounts = {}
            for r in records:
                dt = to_local_datetime(r["created_at"])
                day_key = dt.strftime("%m-%d")
                amount = float(r.get("amount", 0))
                daily_amounts[day_key] = daily_amounts.get(day_key, 0) + amount
            sorted_days = sorted(daily_amounts.keys())
            trend_labels = sorted_days
            trend_amounts = [daily_amounts[d] for d in sorted_days]
        elif year:
            # 按月统计
            monthly_amounts = {}
            for r in records:
                dt = to_local_datetime(r["created_at"])
                month_key = dt.strftime("%Y-%m")
                amount = float(r.get("amount", 0))
                monthly_amounts[month_key] = monthly_amounts.get(month_key, 0) + amount
            sorted_months = sorted(monthly_amounts.keys())
            trend_labels = sorted_months
            trend_amounts = [monthly_amounts[m] for m in sorted_months]
        else:
            # 默认：近12个月
            monthly_amounts = {}
            for r in records:
                dt = to_local_datetime(r["created_at"])
                month_key = dt.strftime("%Y-%m")
                amount = float(r.get("amount", 0))
                monthly_amounts[month_key] = monthly_amounts.get(month_key, 0) + amount
            sorted_months = sorted(monthly_amounts.keys())[-12:]
            trend_labels = sorted_months
            trend_amounts = [monthly_amounts[m] for m in sorted_months]
        
        return {
            "success": True,
            "category_labels": category_labels,
            "category_amounts": category_amounts_list,
            "month_labels": trend_labels,
            "month_amounts": trend_amounts,
            "day_labels": trend_labels if (date or (year and month) or week) else None
        }
    except Exception as e:
        print(f"统计错误: {str(e)[:100]}")
        return {"success": False, "error": str(e)}


@app.get("/api/admin/categories")
async def admin_categories(payload: dict = Depends(verify_admin_token)):
    """分类列表（含记录中的分类 + 手动添加的预设，预设无记录时 count/amount 为 0）"""
    try:
        records = get_records()
        category_stats = {}
        for r in records:
            cat = r.get("category", "其他")
            amount = float(r.get("amount", 0))
            if cat not in category_stats:
                category_stats[cat] = {"count": 0, "amount": 0}
            category_stats[cat]["count"] += 1
            category_stats[cat]["amount"] += amount
        for p in get_category_presets():
            if p and p not in category_stats:
                category_stats[p] = {"count": 0, "amount": 0}
        categories = [{"category": cat, "count": stats["count"], "amount": stats["amount"]} for cat, stats in sorted(category_stats.items())]
        paths = list(category_stats.keys())
        tree = paths_to_tree(paths) if paths else {}
        return {"success": True, "categories": categories, "tree": tree, "paths": sorted(paths)}
    except Exception as e:
        print(f"分类列表错误: {str(e)[:100]}")
        return {"success": False, "error": str(e)}


@app.get("/api/admin/category-tree")
async def admin_get_category_tree(payload: dict = Depends(verify_admin_token)):
    """获取三级类目树（用于级联选择与树形展示）"""
    paths = get_category_tree_paths()
    if paths is None or len(paths) == 0:
        return {"success": True, "tree": {}, "paths": [], "enabled": False}
    return {"success": True, "tree": paths_to_tree(paths), "paths": sorted(paths), "enabled": True}


@app.post("/api/admin/category-tree")
async def admin_set_category_tree(request: Request, payload: dict = Depends(verify_admin_token)):
    """保存三级类目树；body: { "paths": ["正餐", "正餐|早饭", "正餐|午饭|外卖", ...] }"""
    try:
        data = await request.json()
        paths = data.get("paths", [])
        if not isinstance(paths, list):
            return {"success": False, "error": "paths 需为数组"}
        paths = [str(p).strip() for p in paths if str(p).strip()]
        set_category_tree(paths)
        return {"success": True, "paths": sorted(paths)}
    except Exception as e:
        print(f"保存类目树错误: {str(e)[:100]}")
        return {"success": False, "error": str(e)}


@app.post("/api/admin/categories/merge")
async def admin_merge_categories(request: Request, payload: dict = Depends(verify_admin_token)):
    """将旧分类合并到新路径（批量更新记录与别名）；body: { "mappings": [ {"from": "早饭", "to": "正餐|早饭"}, ... ] }"""
    try:
        data = await request.json()
        mappings = data.get("mappings", [])
        if not isinstance(mappings, list) or not mappings:
            return {"success": False, "error": "请提供 mappings 数组"}
        result = merge_categories_to_tree(mappings)
        return {"success": True, "updated": result["success"], "failed": result["failed"], "errors": result.get("errors", [])}
    except Exception as e:
        print(f"合并分类错误: {str(e)[:100]}")
        return {"success": False, "error": str(e)}


@app.post("/api/admin/categories/merge_simple")
async def admin_merge_categories_simple(request: Request, payload: dict = Depends(verify_admin_token)):
    """把多个分散的分类合并成一个。body: { "merge_from": ["早饭","晚饭"], "merge_to": "正餐" }，记录和映射都会更新过去。"""
    try:
        data = await request.json()
        merge_from = data.get("merge_from", [])
        merge_to = (data.get("merge_to") or "").strip()
        if not merge_to:
            return {"success": False, "error": "请填写「合并为」的目标分类名"}
        if not isinstance(merge_from, list) or len(merge_from) == 0:
            return {"success": False, "error": "请至少选择一个要合并的分类"}
        merge_from = [str(x).strip() for x in merge_from if str(x).strip() and str(x).strip() != merge_to]
        if not merge_from:
            return {"success": False, "error": "要合并的分类不能为空且不能与目标相同"}
        mappings = [{"from": name, "to": merge_to} for name in merge_from]
        result = merge_categories_to_tree(mappings)
        return {"success": True, "updated": result["success"], "message": f"已将 {len(merge_from)} 个分类合并为「{merge_to}」，共更新 {result['success']} 条记录"}
    except Exception as e:
        print(f"合并分类错误: {str(e)[:100]}")
        return {"success": False, "error": str(e)}


@app.post("/api/admin/categories/rename")
async def admin_rename_category(
    request: Request,
    payload: dict = Depends(verify_admin_token)
):
    """重命名分类（或删除时转移：将该分类下记录与别名合并到目标）"""
    try:
        data = await request.json()
        old_name = data.get("old_name", "")
        new_name = data.get("new_name", "")
        
        result = rename_category(old_name, new_name)
        if result.get("success"):
            return {"success": True, "count": result.get("count", 0)}
        else:
            return {"success": False, "error": result.get("error", "重命名失败")}
    except Exception as e:
        print(f"重命名分类错误: {str(e)[:100]}")
        return {"success": False, "error": str(e)}


@app.post("/api/admin/categories/add_preset")
async def admin_add_category_preset(request: Request, payload: dict = Depends(verify_admin_token)):
    """网页端新增类目（可多级，如 正餐|晚餐|外卖）。body: { "path": "正餐|晚餐|外卖" }"""
    try:
        data = await request.json()
        path = (data.get("path") or "").strip()
        if not path:
            return {"success": False, "error": "请填写类目路径"}
        add_category_preset(path)
        return {"success": True, "path": path}
    except Exception as e:
        print(f"添加类目预设错误: {str(e)[:100]}")
        return {"success": False, "error": str(e)}


@app.post("/api/admin/categories/remove_preset")
async def admin_remove_category_preset(request: Request, payload: dict = Depends(verify_admin_token)):
    """移除类目预设（仅从预设列表移除，不影响已有记录）"""
    try:
        data = await request.json()
        path = (data.get("path") or "").strip()
        remove_category_preset(path)
        return {"success": True}
    except Exception as e:
        return {"success": False, "error": str(e)}


@app.post("/api/admin/categories/clear_aliases")
async def admin_clear_category_aliases(request: Request, payload: dict = Depends(verify_admin_token)):
    """删除分类（仅清除该分类的别名，用于该分类下无记录时）"""
    try:
        data = await request.json()
        name = (data.get("name") or "").strip()
        if not name:
            return {"success": False, "error": "请指定分类名"}
        ok = clear_category_aliases(name)
        return {"success": ok}
    except Exception as e:
        print(f"清除分类别名错误: {str(e)[:100]}")
        return {"success": False, "error": str(e)}


@app.get("/api/admin/debts")
async def admin_list_debts(
    request: Request,
    payload: dict = Depends(verify_admin_token)
):
    """外债列表（含已还清可选）"""
    try:
        include_paid = request.query_params.get("all") == "1"
        debts = list_debts_all(include_paid=include_paid)
        total = sum(float(d.get("amount", 0)) for d in debts if d.get("status") == "active")
        return {"success": True, "debts": debts, "total_active": total}
    except Exception as e:
        print(f"外债列表错误: {str(e)[:100]}")
        return {"success": False, "error": str(e)}


@app.post("/api/admin/debts")
async def admin_add_debt(
    request: Request,
    payload: dict = Depends(verify_admin_token)
):
    """新增或累加外债"""
    try:
        data = await request.json()
        name = (data.get("name") or "").strip()
        amount = float(data.get("amount", 0))
        note = (data.get("note") or "").strip()
        if not name:
            return {"success": False, "error": "请输入对方姓名/称呼"}
        if amount <= 0:
            return {"success": False, "error": "金额需大于 0"}
        new_amount = add_debt(name, amount, note)
        return {"success": True, "total": new_amount}
    except Exception as e:
        print(f"外债添加错误: {str(e)[:100]}")
        return {"success": False, "error": str(e)}


@app.post("/api/admin/debts/repay")
async def admin_repay_debt(
    request: Request,
    payload: dict = Depends(verify_admin_token)
):
    """还款"""
    try:
        data = await request.json()
        name = (data.get("name") or "").strip()
        amount = float(data.get("amount", 0))
        if not name:
            return {"success": False, "error": "请输入对方姓名"}
        if amount <= 0:
            return {"success": False, "error": "金额需大于 0"}
        result = repay_debt(name, amount)
        if result.get("error") == "not_found":
            return {"success": False, "error": f"未找到欠 {name} 的记录"}
        if result.get("error") == "overpay":
            return {"success": False, "error": f"当前欠 {name} {result.get('balance', 0):.2f} 元，还款金额超出"}
        return {"success": True, "balance": result.get("balance", 0), "status": result.get("status", "active")}
    except Exception as e:
        print(f"外债还款错误: {str(e)[:100]}")
        return {"success": False, "error": str(e)}


@app.delete("/api/admin/debts")
async def admin_delete_debt(
    request: Request,
    payload: dict = Depends(verify_admin_token)
):
    """删除外债记录（整条清空）"""
    try:
        name = (request.query_params.get("name") or "").strip()
        if not name:
            return {"success": False, "error": "请指定 name"}
        ok = delete_debt(name)
        return {"success": ok}
    except Exception as e:
        print(f"外债删除错误: {str(e)[:100]}")
        return {"success": False, "error": str(e)}


@app.get("/api/admin/export_categories")
async def admin_export_categories(payload: dict = Depends(verify_admin_token)):
    """管理后台：下载分类表 Excel（与 export_categories 相同，需登录）"""
    try:
        data = build_category_excel_bytes()
        filename = "categories.xlsx"
        from urllib.parse import quote
        encoded_filename = quote(filename)
        headers = {
            "Content-Disposition": f'attachment; filename="{filename}"; filename*=UTF-8\'\'{encoded_filename}',
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        }
        return Response(content=data, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)
    except Exception as e:
        print(f"导出分类错误: {str(e)[:100]}")
        return Response(content="error", status_code=500)


@app.get("/api/admin/export_category_mapping")
async def admin_export_category_mapping(payload: dict = Depends(verify_admin_token)):
    """下载分类映射表模板（含当前所有类目），编辑后上传用于合并；删掉的行上传后为待处理"""
    try:
        data = build_category_mapping_excel_bytes()
        headers = {
            "Content-Disposition": 'attachment; filename="category_mapping.xlsx"',
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        }
        return Response(content=data, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers=headers)
    except Exception as e:
        print(f"导出映射表错误: {str(e)[:200]}")
        return Response(content="error", status_code=500)


@app.post("/api/admin/import_category_mapping")
async def admin_import_category_mapping(file: UploadFile = File(...), payload: dict = Depends(verify_admin_token)):
    """上传分类映射表：按表合并；表中被删掉的原分类其记录列为待处理，返回待处理列表供前端按组设置"""
    try:
        file_bytes = await file.read()
        parsed = parse_category_mapping_excel(file_bytes)
        if parsed.get("error"):
            return {"success": False, "error": parsed["error"], "deleted": []}
        mappings = parsed.get("mappings", [])
        origins_in_file = set(parsed.get("origins_in_file", []))
        stats_before = {x["category"]: x["count"] for x in get_category_stats()}
        deleted = [{"category": cat, "count": stats_before.get(cat, 0)} for cat in stats_before if cat not in origins_in_file]
        to_merge = [m for m in mappings if (m.get("from") or "").strip() != (m.get("to") or "").strip()]
        updated = 0
        if to_merge:
            result = merge_categories_to_tree(to_merge)
            updated = result.get("success", 0)
        return {"success": True, "updated": updated, "deleted": deleted, "message": f"已合并更新 {updated} 条记录" + ("；以下类目已从表中删除，请为每组设置新分类" if deleted else "")}
    except Exception as e:
        print(f"导入映射表错误: {str(e)[:100]}")
        return {"success": False, "error": str(e), "deleted": []}


@app.post("/api/admin/import_categories")
async def admin_import_categories(file: UploadFile = File(...), payload: dict = Depends(verify_admin_token)):
    """管理后台：上传分类表 Excel 批量重命名"""
    try:
        file_bytes = await file.read()
        parsed = parse_category_excel(file_bytes)
        if parsed.get("error"):
            return Response(content=parsed["error"], status_code=400)
        renames = parsed.get("renames", [])
        if not renames:
            return Response(content="no_changes", status_code=400)
        result = batch_rename_categories(renames)
        return {"success": True, "renamed": result["success"], "failed": result["failed"], "message": f"已重命名 {result['success']} 个分类，失败 {len(result['failed'])} 个"}
    except Exception as e:
        print(f"导入分类错误: {str(e)[:100]}")
        return {"success": False, "error": str(e)}


@app.post("/api/admin/records/batch_set_category")
async def admin_batch_set_category(request: Request, payload: dict = Depends(verify_admin_token)):
    """批量修改选中记录的分类"""
    try:
        data = await request.json()
        ids = data.get("ids", [])
        category = (data.get("category") or "").strip()
        if not ids or not category:
            return {"success": False, "error": "请选择记录并指定分类"}
        supabase = get_supabase_client()
        updated = 0
        for rid in ids:
            try:
                r = supabase.table("records").select("amount,description").eq("id", rid).execute()
                if not r.data or len(r.data) == 0:
                    continue
                row = r.data[0]
                supabase.table("records").update({"category": category}).eq("id", rid).execute()
                updated += 1
            except Exception:
                pass
        invalidate_records_cache()
        return {"success": True, "updated": updated}
    except Exception as e:
        print(f"批量改分类错误: {str(e)[:100]}")
        return {"success": False, "error": str(e)}


@app.get("/api/admin/settings")
async def admin_get_settings(payload: dict = Depends(verify_admin_token)):
    """获取所有设置"""
    try:
        supabase = get_supabase_client()
        result = supabase.table("settings").select("key,value").execute()
        settings = {}
        for item in result.data:
            settings[item["key"]] = item["value"]
        return {"success": True, "settings": settings}
    except Exception as e:
        print(f"获取设置错误: {str(e)[:100]}")
        return {"success": True, "settings": {}}


@app.post("/api/admin/settings")
async def admin_save_settings(
    request: Request,
    payload: dict = Depends(verify_admin_token)
):
    """保存设置"""
    try:
        data = await request.json()
        key = data.get("key", "")
        value = data.get("value", "")
        
        if not key:
            return {"success": False, "error": "设置项名称不能为空"}
        
        if set_setting(key, value):
            return {"success": True}
        else:
            return {"success": False, "error": "保存失败"}
    except Exception as e:
        print(f"保存设置错误: {str(e)[:100]}")
        return {"success": False, "error": str(e)}


@app.get("/api/admin/monthly_stats")
async def admin_monthly_stats(
    request: Request,
    payload: dict = Depends(verify_admin_token)
):
    """获取今年所有月份的统计。?fresh=1 时强制从数据库重新加载。"""
    try:
        params = dict(request.query_params)
        force_refresh = params.get("fresh") == "1"
        now = datetime.now(LOCAL_TZ)
        year = now.year
        all_records = get_records_cached(force_refresh=force_refresh)
        print(f"月度统计: 获取到 {len(all_records)} 条记录")
        
        monthly_stats = {}
        for month in range(1, 13):
            month_start = datetime(year, month, 1, 0, 0, 0, tzinfo=LOCAL_TZ)
            if month == 12:
                month_end = datetime(year + 1, 1, 1, 0, 0, 0, tzinfo=LOCAL_TZ)
            else:
                month_end = datetime(year, month + 1, 1, 0, 0, 0, tzinfo=LOCAL_TZ)
            
            month_records = filter_records_by_local_range(all_records, month_start, month_end)
            month_amount = sum(float(r["amount"]) for r in month_records)
            month_count = len(month_records)
            
            monthly_stats[f"{year}-{month:02d}"] = {
                "month": month,
                "amount": month_amount,
                "count": month_count
            }
        
        return {
            "success": True,
            "year": year,
            "months": monthly_stats,
            "total_records": len(all_records)
        }
    except Exception as e:
        import traceback
        error_detail = traceback.format_exc()
        print(f"月度统计错误: {error_detail}")
        return {"success": False, "error": str(e)}


@app.get("/api/admin/daily_stats")
async def admin_daily_stats(
    request: Request,
    payload: dict = Depends(verify_admin_token)
):
    """获取指定月份的每日统计。?fresh=1 时强制从数据库重新加载。"""
    try:
        params = dict(request.query_params)
        force_refresh = params.get("fresh") == "1"
        year = int(params.get("year", datetime.now(LOCAL_TZ).year))
        month = int(params.get("month", datetime.now(LOCAL_TZ).month))
        month_start = datetime(year, month, 1, 0, 0, 0, tzinfo=LOCAL_TZ)
        if month == 12:
            month_end = datetime(year + 1, 1, 1, 0, 0, 0, tzinfo=LOCAL_TZ)
        else:
            month_end = datetime(year, month + 1, 1, 0, 0, 0, tzinfo=LOCAL_TZ)
        all_records = get_records_cached(force_refresh=force_refresh)
        month_records = filter_records_by_local_range(all_records, month_start, month_end)
        
        daily_stats = {}
        for r in month_records:
            dt = to_local_datetime(r["created_at"])
            day_key = dt.strftime("%Y-%m-%d")
            amount = float(r["amount"])
            if day_key not in daily_stats:
                daily_stats[day_key] = {"amount": 0, "count": 0}
            daily_stats[day_key]["amount"] += amount
            daily_stats[day_key]["count"] += 1
        
        # 生成该月所有日期（即使没有记录）
        days = []
        current = month_start
        while current < month_end:
            day_key = current.strftime("%Y-%m-%d")
            days.append({
                "date": day_key,
                "day": current.day,
                "amount": daily_stats.get(day_key, {}).get("amount", 0),
                "count": daily_stats.get(day_key, {}).get("count", 0)
            })
            current += timedelta(days=1)
        
        return {
            "success": True,
            "year": year,
            "month": month,
            "days": days
        }
    except Exception as e:
        print(f"每日统计错误: {str(e)[:100]}")
        return {"success": False, "error": str(e)}


@app.get("/api/admin/date_records")
async def admin_date_records(
    request: Request,
    payload: dict = Depends(verify_admin_token)
):
    """获取指定日期的明细记录"""
    try:
        params = dict(request.query_params)
        date_str = params.get("date", "")
        
        if not date_str:
            return {"success": False, "error": "缺少日期参数"}
        
        date_obj = datetime.strptime(date_str, "%Y-%m-%d").replace(tzinfo=LOCAL_TZ)
        date_start = date_obj
        date_end = date_obj + timedelta(days=1)
        
        all_records = get_records_cached()
        date_records = filter_records_by_local_range(all_records, date_start, date_end)
        
        formatted = []
        for r in date_records:
            dt = to_local_datetime(r["created_at"])
            formatted.append({
                "id": r["id"],
                "time": dt.strftime("%H:%M"),
                "description": r.get("description", ""),
                "amount": float(r.get("amount", 0)),
                "category": r.get("category", "")
            })
        
        # 按时间排序
        formatted.sort(key=lambda x: x["time"])
        
        return {
            "success": True,
            "date": date_str,
            "records": formatted,
            "total": sum(r["amount"] for r in formatted)
        }
    except Exception as e:
        print(f"日期明细错误: {str(e)[:100]}")
        return {"success": False, "error": str(e)}


@app.get("/api/admin/month_category_stats")
async def admin_month_category_stats(
    request: Request,
    payload: dict = Depends(verify_admin_token)
):
    """获取指定月份的分类统计"""
    try:
        params = dict(request.query_params)
        year = int(params.get("year", datetime.now(LOCAL_TZ).year))
        month = int(params.get("month", datetime.now(LOCAL_TZ).month))
        
        month_start = datetime(year, month, 1, 0, 0, 0, tzinfo=LOCAL_TZ)
        if month == 12:
            month_end = datetime(year + 1, 1, 1, 0, 0, 0, tzinfo=LOCAL_TZ)
        else:
            month_end = datetime(year, month + 1, 1, 0, 0, 0, tzinfo=LOCAL_TZ)
        
        all_records = get_records_cached()
        month_records = filter_records_by_local_range(all_records, month_start, month_end)
        
        # 按分类统计
        category_stats = {}
        for r in month_records:
            category = r.get("category", "未分类")
            amount = float(r.get("amount", 0))
            if category not in category_stats:
                category_stats[category] = {"amount": 0, "count": 0}
            category_stats[category]["amount"] += amount
            category_stats[category]["count"] += 1
        
        # 转换为列表并按金额排序
        category_list = [
            {
                "category": cat,
                "amount": stats["amount"],
                "count": stats["count"]
            }
            for cat, stats in category_stats.items()
        ]
        category_list.sort(key=lambda x: x["amount"], reverse=True)
        
        # 前10名
        top10 = category_list[:10]
        
        return {
            "success": True,
            "year": year,
            "month": month,
            "top10": top10,
            "all": category_list,
            "total": sum(c["amount"] for c in category_list)
        }
    except Exception as e:
        print(f"月份分类统计错误: {str(e)[:100]}")
        return {"success": False, "error": str(e)}


@app.get("/api/admin/year_category_stats")
async def admin_year_category_stats(
    request: Request,
    payload: dict = Depends(verify_admin_token)
):
    """获取指定年份的分类统计"""
    try:
        params = dict(request.query_params)
        year = int(params.get("year", datetime.now(LOCAL_TZ).year))
        
        year_start = datetime(year, 1, 1, 0, 0, 0, tzinfo=LOCAL_TZ)
        year_end = datetime(year + 1, 1, 1, 0, 0, 0, tzinfo=LOCAL_TZ)
        
        all_records = get_records_cached()
        year_records = filter_records_by_local_range(all_records, year_start, year_end)
        
        # 按分类统计
        category_stats = {}
        for r in year_records:
            category = r.get("category", "未分类")
            amount = float(r.get("amount", 0))
            if category not in category_stats:
                category_stats[category] = {"amount": 0, "count": 0}
            category_stats[category]["amount"] += amount
            category_stats[category]["count"] += 1
        
        # 转换为列表并按金额排序
        category_list = [
            {
                "category": cat,
                "amount": stats["amount"],
                "count": stats["count"]
            }
            for cat, stats in category_stats.items()
        ]
        category_list.sort(key=lambda x: x["amount"], reverse=True)
        
        return {
            "success": True,
            "year": year,
            "categories": category_list,
            "total": sum(c["amount"] for c in category_list)
        }
    except Exception as e:
        print(f"年份分类统计错误: {str(e)[:100]}")
        return {"success": False, "error": str(e)}


@app.get("/api/admin/date_category_stats")
async def admin_date_category_stats(
    request: Request,
    payload: dict = Depends(verify_admin_token)
):
    """获取指定日期的分类统计"""
    try:
        params = dict(request.query_params)
        date_str = params.get("date", "")
        
        if not date_str:
            return {"success": False, "error": "缺少日期参数"}
        
        date_obj = datetime.strptime(date_str, "%Y-%m-%d").replace(tzinfo=LOCAL_TZ)
        date_start = date_obj
        date_end = date_obj + timedelta(days=1)
        
        all_records = get_records_cached()
        date_records = filter_records_by_local_range(all_records, date_start, date_end)
        
        # 按分类统计
        category_stats = {}
        for r in date_records:
            category = r.get("category", "未分类")
            amount = float(r.get("amount", 0))
            if category not in category_stats:
                category_stats[category] = {"amount": 0, "count": 0}
            category_stats[category]["amount"] += amount
            category_stats[category]["count"] += 1
        
        # 转换为列表并按金额排序
        category_list = [
            {
                "category": cat,
                "amount": stats["amount"],
                "count": stats["count"]
            }
            for cat, stats in category_stats.items()
        ]
        category_list.sort(key=lambda x: x["amount"], reverse=True)
        
        return {
            "success": True,
            "date": date_str,
            "categories": category_list,
            "total": sum(c["amount"] for c in category_list)
        }
    except Exception as e:
        print(f"日期分类统计错误: {str(e)[:100]}")
        return {"success": False, "error": str(e)}


@app.get("/api/admin/category_records")
async def admin_category_records(
    request: Request,
    payload: dict = Depends(verify_admin_token)
):
    """获取指定分类在指定时间范围内的明细记录"""
    try:
        params = dict(request.query_params)
        category = params.get("category", "")
        year = params.get("year", "")
        month = params.get("month", "")
        date = params.get("date", "")
        
        if not category:
            return {"success": False, "error": "缺少分类参数"}
        
        all_records = get_records_cached()
        
        # 根据时间范围筛选
        if date:
            # 单日
            date_obj = datetime.strptime(date, "%Y-%m-%d").replace(tzinfo=LOCAL_TZ)
            date_start = date_obj
            date_end = date_obj + timedelta(days=1)
            filtered = filter_records_by_local_range(all_records, date_start, date_end)
        elif year and month:
            # 单月
            year_int = int(year)
            month_int = int(month)
            month_start = datetime(year_int, month_int, 1, 0, 0, 0, tzinfo=LOCAL_TZ)
            if month_int == 12:
                month_end = datetime(year_int + 1, 1, 1, 0, 0, 0, tzinfo=LOCAL_TZ)
            else:
                month_end = datetime(year_int, month_int + 1, 1, 0, 0, 0, tzinfo=LOCAL_TZ)
            filtered = filter_records_by_local_range(all_records, month_start, month_end)
        elif year:
            # 全年
            year_int = int(year)
            year_start = datetime(year_int, 1, 1, 0, 0, 0, tzinfo=LOCAL_TZ)
            year_end = datetime(year_int + 1, 1, 1, 0, 0, 0, tzinfo=LOCAL_TZ)
            filtered = filter_records_by_local_range(all_records, year_start, year_end)
        else:
            filtered = all_records
        
        # 按分类筛选
        category_records = [r for r in filtered if r.get("category", "") == category]
        
        # 格式化
        formatted = []
        for r in category_records:
            dt = to_local_datetime(r["created_at"])
            formatted.append({
                "id": r["id"],
                "date": dt.strftime("%Y-%m-%d"),
                "time": dt.strftime("%H:%M"),
                "description": r.get("description", ""),
                "amount": float(r.get("amount", 0)),
                "category": r.get("category", "")
            })
        
        # 按日期和时间排序
        formatted.sort(key=lambda x: (x["date"], x["time"]))
        
        return {
            "success": True,
            "category": category,
            "records": formatted,
            "total": sum(r["amount"] for r in formatted),
            "count": len(formatted)
        }
    except Exception as e:
        print(f"分类明细错误: {str(e)[:100]}")
        return {"success": False, "error": str(e)}


@app.post("/api/admin/records/batch_delete")
async def admin_batch_delete_records(
    request: Request,
    payload: dict = Depends(verify_admin_token)
):
    """批量删除记录"""
    try:
        data = await request.json()
        record_ids = data.get("ids", [])
        
        if not record_ids:
            return {"success": False, "error": "请选择要删除的记录"}
        
        deleted_count = 0
        for record_id in record_ids:
            if delete_record(record_id):
                deleted_count += 1
        
        return {
            "success": True,
            "deleted_count": deleted_count,
            "total": len(record_ids)
        }
    except Exception as e:
        print(f"批量删除错误: {str(e)[:100]}")
        return {"success": False, "error": str(e)}


def verify_admin_token_flexible(request: Request):
    """验证管理员token（支持header和query参数）"""
    try:
        # 先尝试从header获取
        auth_header = request.headers.get("Authorization", "")
        token = None
        if auth_header.startswith("Bearer "):
            token = auth_header[7:]
        else:
            # 从query参数获取
            params = dict(request.query_params)
            token = params.get("token", "")
        
        if not token:
            raise HTTPException(status_code=401, detail="Missing token")
        
        payload = jwt.decode(token, ADMIN_SECRET, algorithms=["HS256"])
        if payload.get("type") != "admin":
            raise HTTPException(status_code=403, detail="Invalid token type")
        return payload
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=403, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=403, detail="Invalid token")


@app.get("/api/admin/export")
async def admin_export(
    request: Request,
    payload: dict = Depends(verify_admin_token_flexible)
):
    """管理后台导出Excel"""
    try:
        params = dict(request.query_params)
        period = params.get("period", "all")  # all, month, year, custom
        year = params.get("year", "")
        month = params.get("month", "")
        date_from = params.get("date_from", "")
        date_to = params.get("date_to", "")
        
        all_records = get_records_cached()
        
        # 根据period筛选记录
        if period == "all":
            filtered = all_records
        elif period == "month" and year and month:
            year_int = int(year)
            month_int = int(month)
            month_start = datetime(year_int, month_int, 1, 0, 0, 0, tzinfo=LOCAL_TZ)
            if month_int == 12:
                month_end = datetime(year_int + 1, 1, 1, 0, 0, 0, tzinfo=LOCAL_TZ)
            else:
                month_end = datetime(year_int, month_int + 1, 1, 0, 0, 0, tzinfo=LOCAL_TZ)
            filtered = filter_records_by_local_range(all_records, month_start, month_end)
        elif period == "year" and year:
            year_int = int(year)
            year_start = datetime(year_int, 1, 1, 0, 0, 0, tzinfo=LOCAL_TZ)
            year_end = datetime(year_int + 1, 1, 1, 0, 0, 0, tzinfo=LOCAL_TZ)
            filtered = filter_records_by_local_range(all_records, year_start, year_end)
        elif period == "custom" and date_from and date_to:
            start_date = datetime.strptime(date_from, "%Y-%m-%d").replace(tzinfo=LOCAL_TZ)
            end_date = datetime.strptime(date_to, "%Y-%m-%d").replace(tzinfo=LOCAL_TZ) + timedelta(days=1)
            filtered = filter_records_by_local_range(all_records, start_date, end_date)
        else:
            filtered = all_records

        # 按分类筛选（可选）
        categories_param = params.get("categories", "")
        if categories_param:
            cat_list = [c.strip() for c in categories_param.split(",") if c.strip()]
            if cat_list:
                filtered = [r for r in filtered if r.get("category", "").strip() in cat_list]
        
        # 导出用时间范围（用于 Excel 表头）
        if filtered:
            from_dates = [to_local_datetime(r["created_at"]) for r in filtered]
            export_start = min(from_dates).replace(hour=0, minute=0, second=0, microsecond=0)
            export_end = max(from_dates) + timedelta(days=1)
        else:
            export_start = datetime.now(LOCAL_TZ).replace(hour=0, minute=0, second=0, microsecond=0)
            export_end = export_start + timedelta(days=1)
        
        # 生成Excel
        excel_bytes = build_export_excel_bytes(filtered, export_start, export_end, limit=10000)
        
        # 生成文件名
        now = datetime.now(LOCAL_TZ)
        if period == "month" and year and month:
            filename = f"records_{year}{month:02d}_export.xlsx"
        elif period == "year" and year:
            filename = f"records_{year}_export.xlsx"
        elif period == "custom":
            filename = f"records_{date_from}_to_{date_to}_export.xlsx"
        else:
            filename = f"records_all_{now.strftime('%Y%m%d_%H%M%S')}_export.xlsx"
        
        return Response(
            content=excel_bytes,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        print(f"导出错误: {str(e)[:100]}")
        return {"success": False, "error": str(e)}


@app.get("/api/admin/backup")
async def admin_backup(
    request: Request,
    payload: dict = Depends(verify_admin_token_flexible)
):
    """数据备份（导出所有数据）"""
    try:
        all_records = get_records_cached()
        excel_bytes = build_export_excel_bytes(all_records)
        
        now = datetime.now(LOCAL_TZ)
        filename = f"backup_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return Response(
            content=excel_bytes,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    except Exception as e:
        print(f"备份错误: {str(e)[:100]}")
        return {"success": False, "error": str(e)}


@app.get("/api/admin/comparison")
async def admin_comparison(
    request: Request,
    payload: dict = Depends(verify_admin_token)
):
    """数据对比分析"""
    try:
        params = dict(request.query_params)
        type = params.get("type", "month")  # month, year
        
        now = datetime.now(LOCAL_TZ)
        all_records = get_records_cached()
        
        if type == "month":
            # 本月 vs 上月
            current_month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            if now.month == 1:
                last_month_start = datetime(now.year - 1, 12, 1, 0, 0, 0, tzinfo=LOCAL_TZ)
                last_month_end = current_month_start
            else:
                last_month_start = datetime(now.year, now.month - 1, 1, 0, 0, 0, tzinfo=LOCAL_TZ)
                last_month_end = current_month_start
            
            current_records = filter_records_by_local_range(all_records, current_month_start, now + timedelta(days=1))
            last_records = filter_records_by_local_range(all_records, last_month_start, last_month_end)
            
            current_amount = sum(float(r["amount"]) for r in current_records)
            last_amount = sum(float(r["amount"]) for r in last_records)
            
            return {
                "success": True,
                "type": "month",
                "current": {
                    "period": f"{now.year}年{now.month}月",
                    "amount": current_amount,
                    "count": len(current_records)
                },
                "last": {
                    "period": f"{last_month_start.year}年{last_month_start.month}月",
                    "amount": last_amount,
                    "count": len(last_records)
                },
                "change": current_amount - last_amount,
                "change_percent": ((current_amount - last_amount) / last_amount * 100) if last_amount > 0 else 0
            }
        elif type == "year":
            # 今年 vs 去年
            current_year_start = datetime(now.year, 1, 1, 0, 0, 0, tzinfo=LOCAL_TZ)
            last_year_start = datetime(now.year - 1, 1, 1, 0, 0, 0, tzinfo=LOCAL_TZ)
            last_year_end = current_year_start
            
            current_records = filter_records_by_local_range(all_records, current_year_start, now + timedelta(days=1))
            last_records = filter_records_by_local_range(all_records, last_year_start, last_year_end)
            
            current_amount = sum(float(r["amount"]) for r in current_records)
            last_amount = sum(float(r["amount"]) for r in last_records)
            
            return {
                "success": True,
                "type": "year",
                "current": {
                    "period": f"{now.year}年",
                    "amount": current_amount,
                    "count": len(current_records)
                },
                "last": {
                    "period": f"{now.year - 1}年",
                    "amount": last_amount,
                    "count": len(last_records)
                },
                "change": current_amount - last_amount,
                "change_percent": ((current_amount - last_amount) / last_amount * 100) if last_amount > 0 else 0
            }
        else:
            return {"success": False, "error": "不支持的对比类型"}
    except Exception as e:
        print(f"对比分析错误: {str(e)[:100]}")
        return {"success": False, "error": str(e)}


@app.get("/api/admin/weekly_stats")
async def admin_weekly_stats(payload: dict = Depends(verify_admin_token)):
    """周统计"""
    try:
        now = datetime.now(LOCAL_TZ)
        # 本周一
        days_since_monday = now.weekday()
        week_start = (now - timedelta(days=days_since_monday)).replace(hour=0, minute=0, second=0, microsecond=0)
        week_end = now + timedelta(days=1)
        
        all_records = get_records_cached()
        week_records = filter_records_by_local_range(all_records, week_start, week_end)
        
        # 按天统计
        daily_stats = {}
        for r in week_records:
            dt = to_local_datetime(r["created_at"])
            day_key = dt.strftime("%Y-%m-%d")
            amount = float(r["amount"])
            if day_key not in daily_stats:
                daily_stats[day_key] = {"amount": 0, "count": 0, "date": day_key}
            daily_stats[day_key]["amount"] += amount
            daily_stats[day_key]["count"] += 1
        
        # 生成本周所有日期
        days = []
        current = week_start
        while current < week_end:
            day_key = current.strftime("%Y-%m-%d")
            days.append({
                "date": day_key,
                "day": current.day,
                "weekday": ["周一", "周二", "周三", "周四", "周五", "周六", "周日"][current.weekday()],
                "amount": daily_stats.get(day_key, {}).get("amount", 0),
                "count": daily_stats.get(day_key, {}).get("count", 0)
            })
            current += timedelta(days=1)
        
        total_amount = sum(float(r["amount"]) for r in week_records)
        avg_daily = total_amount / 7 if len(days) > 0 else 0
        
        return {
            "success": True,
            "week_start": week_start.strftime("%Y-%m-%d"),
            "week_end": (week_end - timedelta(days=1)).strftime("%Y-%m-%d"),
            "days": days,
            "total_amount": total_amount,
            "total_count": len(week_records),
            "avg_daily": avg_daily
        }
    except Exception as e:
        print(f"周统计错误: {str(e)[:100]}")
        return {"success": False, "error": str(e)}


@app.get("/api/admin/quarterly_stats")
async def admin_quarterly_stats(
    request: Request,
    payload: dict = Depends(verify_admin_token)
):
    """季度统计"""
    try:
        params = dict(request.query_params)
        year = int(params.get("year", datetime.now(LOCAL_TZ).year))
        
        quarters = []
        for q in range(1, 5):
            month_start = (q - 1) * 3 + 1
            quarter_start = datetime(year, month_start, 1, 0, 0, 0, tzinfo=LOCAL_TZ)
            if q == 4:
                quarter_end = datetime(year + 1, 1, 1, 0, 0, 0, tzinfo=LOCAL_TZ)
            else:
                quarter_end = datetime(year, month_start + 3, 1, 0, 0, 0, tzinfo=LOCAL_TZ)
            
            all_records = get_records_cached()
            quarter_records = filter_records_by_local_range(all_records, quarter_start, quarter_end)
            
            quarters.append({
                "quarter": q,
                "period": f"Q{q}",
                "amount": sum(float(r["amount"]) for r in quarter_records),
                "count": len(quarter_records)
            })
        
        return {
            "success": True,
            "year": year,
            "quarters": quarters,
            "total": sum(q["amount"] for q in quarters)
        }
    except Exception as e:
        print(f"季度统计错误: {str(e)[:100]}")
        return {"success": False, "error": str(e)}


@app.get("/api/admin/avg_daily")
async def admin_avg_daily(
    request: Request,
    payload: dict = Depends(verify_admin_token)
):
    """平均每日支出"""
    try:
        params = dict(request.query_params)
        period = params.get("period", "month")  # month, year, all
        
        now = datetime.now(LOCAL_TZ)
        all_records = get_records_cached()
        
        if period == "month":
            month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
            filtered = filter_records_by_local_range(all_records, month_start, now + timedelta(days=1))
            days = (now - month_start).days + 1
        elif period == "year":
            year_start = datetime(now.year, 1, 1, 0, 0, 0, tzinfo=LOCAL_TZ)
            filtered = filter_records_by_local_range(all_records, year_start, now + timedelta(days=1))
            days = (now - year_start).days + 1
        else:
            # all
            if not all_records:
                return {
                    "success": True,
                    "period": "all",
                    "avg_daily": 0,
                    "total_amount": 0,
                    "days": 0
                }
            first_record = min(all_records, key=lambda r: to_local_datetime(r["created_at"]))
            first_date = to_local_datetime(first_record["created_at"]).replace(hour=0, minute=0, second=0, microsecond=0)
            filtered = all_records
            days = (now - first_date).days + 1
        
        total_amount = sum(float(r["amount"]) for r in filtered)
        avg_daily = total_amount / days if days > 0 else 0
        
        return {
            "success": True,
            "period": period,
            "avg_daily": avg_daily,
            "total_amount": total_amount,
            "days": days
        }
    except Exception as e:
        print(f"平均每日支出错误: {str(e)[:100]}")
        return {"success": False, "error": str(e)}


